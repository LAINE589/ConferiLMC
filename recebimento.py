"""
Módulo de Acompanhamento de Recebimento de Combustíveis
Confronta notas fiscais de compra (Excel) com o Estoque Físico do DAC (PDF).
"""
import io, re
import xlrd
import pypdf
from datetime import datetime


# ── Normalização de nomes de produtos ─────────────────────────────────────────
MAPA_PRODUTO = [
    # ETANOL — sempre normaliza para ETANOL HIDRATADO COMUM
    (['ETANOL HIDRATADO', 'ETANOL COMUM', 'ETANOL'], 'ETANOL HIDRATADO COMUM'),

    # GASOLINA ADITIVADA — verificar ANTES de gasolina comum para capturar "C ADIT"
    (['GASOLINA COMUM C ADITIVADA', 'GASOLINA COMUM ADITIVADA',
      'GASOLINA COMUM C ADIT', 'GASOLINA C ADITIVADA',
      'GASOLINA ADITIVADA', 'GASOLINA C ADIT', 'GASOLINA ADIT'], 'GASOLINA ADITIVADA'),

    # GASOLINA COMUM
    (['GASOLINA COMUM C', 'GASOLINA COMUM', 'GASOLINA C'], 'GASOLINA COMUM'),

    # DIESEL S500
    (['DIESEL B S500', 'DIESEL S500', 'OLEO DIESEL B S500',
      'ÓLEO DIESEL B S500', 'DIESEL S-500'], 'DIESEL S500'),

    # DIESEL S10 — aditivado e normal mapeiam para o mesmo nome do DAC
    (['DIESEL B S10', 'DIESEL S10', 'OLEO DIESEL B S10',
      'ÓLEO DIESEL B S10', 'DIESEL S-10'], 'DIESEL S10'),

    (['DIESEL'], 'DIESEL'),
]

def normalizar_produto(nome):
    n = nome.upper().strip()
    for chaves, padrao in MAPA_PRODUTO:
        for c in chaves:
            if c in n:
                return padrao
    return nome.strip()


def _fl(v):
    if v is None or str(v).strip() in ('', 'None'): return None
    try: return float(str(v).replace('.','').replace(',','.'))
    except: return None

def _xl_date(val):
    try:
        return datetime.fromordinal(datetime(1900,1,1).toordinal() + int(val) - 2)
    except: return None


# ── Leitor de Notas Fiscais (Excel .xls / .xlsx) ──────────────────────────────
def _converter_xls_para_rows(arquivo_bytes):
    """Converte XLS (qualquer versão) para lista de rows usando LibreOffice como fallback."""
    # Tentar xlrd primeiro (mais rápido)
    try:
        wb = xlrd.open_workbook(file_contents=arquivo_bytes, on_demand=True)
        ws = wb.sheet_by_index(0)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == 3:  # data
                    row.append(_xl_date(cell.value))
                elif cell.ctype == 2:  # número
                    row.append(cell.value)
                else:
                    row.append(str(cell.value).strip())
            rows.append(row)
        return rows
    except Exception:
        pass

    # Fallback: converter via LibreOffice
    import subprocess, tempfile, glob as _glob, os as _os
    tmp = tempfile.NamedTemporaryFile(suffix='.xls', delete=False)
    tmp.write(arquivo_bytes); tmp.close()
    outdir = tempfile.mkdtemp()
    subprocess.run(
        ['python3', '/mnt/skills/public/pptx/scripts/office/soffice.py',
         '--headless', '--convert-to', 'xlsx', '--outdir', outdir, tmp.name],
        capture_output=True, timeout=60
    )
    _os.unlink(tmp.name)
    converted = _glob.glob(_os.path.join(outdir, '*.xlsx'))
    if not converted:
        return None

    import openpyxl
    wb2 = openpyxl.load_workbook(converted[0], data_only=True)
    ws2 = wb2.active
    rows = []
    for row in ws2.iter_rows(values_only=True):
        rows.append([v if v is not None else '' for v in row])
    return rows


def ler_notas_xls(arquivo_bytes, filename):
    """
    Lê a relação de notas fiscais de entrada.
    Detecta formato automaticamente (XLS/XLSX) e usa LibreOffice como fallback.
    """
    notas = []

    # Detectar formato pelo magic bytes
    is_xlsx = arquivo_bytes[:2] == b'PK'
    ext = 'xlsx' if is_xlsx else 'xls'

    try:
        if ext == 'xlsx':
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), data_only=True)
            ws = wb.active
            rows = [[v if v is not None else '' for v in row]
                    for row in ws.iter_rows(values_only=True)]
        else:
            rows = _converter_xls_para_rows(arquivo_bytes)
            if rows is None:
                return None, "Não foi possível ler o arquivo Excel. Tente exportar em outro formato."

        # Detectar colunas pelo cabeçalho
        col_op=0; col_nf=1; col_data=2; col_prod=8; col_qtd=15; col_unit=16

        for row in rows:
            def s(v): return str(v).strip() if v is not None else ''
            vals = [s(v) for v in row]
            if not vals or not vals[0]: continue

            # Metadados
            if vals[0].startswith('Empresa:'): continue
            if vals[0].startswith('CNPJ:'): continue

            # Detectar linha de cabeçalho
            v0 = vals[0].lower()
            if v0 in ('operação','operacao','tipo','op.','operaçao'):
                for ci, val in enumerate(vals):
                    v = val.lower().strip()
                    if v in ('operação','operacao','tipo','op.'): col_op = ci
                    elif v in ('nº nf-e','nf','nf-e','número','numero','nº nf'): col_nf = ci
                    elif v == 'data': col_data = ci
                    elif ('nome' in v and 'produto' in v) or ('produto' in v and 'cod' not in v and 'ncm' not in v and 'valor' not in v and col_prod == 8): col_prod = ci
                    elif v in ('quant.','quantidade','qtd','qtde'): col_qtd = ci
                    elif v in ('fator unit.','unidade','un','unit','und'): col_unit = ci
                continue

            # Linha de dados
            if vals[col_op].lower() != 'entrada': continue

            try:
                prod = s(row[col_prod]) if len(row) > col_prod else ''
                if not prod: continue
                qtd_raw = row[col_qtd] if len(row) > col_qtd else ''
                qtd_val = str(qtd_raw).replace(',','.') if qtd_raw != '' else '0'
                try: qtd = float(qtd_val)
                except: continue
                if qtd <= 0: continue

                unit = s(row[col_unit]).upper() if len(row) > col_unit else ''
                # Aceitar apenas se unidade for L, ou se produto for combustível reconhecido
                prod_norm_test = normalizar_produto(prod)
                eh_combustivel = any(c in prod_norm_test.upper() for c in
                                     ['ETANOL','GASOLINA','DIESEL','GNV'])
                if unit and unit not in ('L','LT','LTR','LITRO','LITROS'): continue
                if not unit and not eh_combustivel: continue

                data_val = row[col_data]
                if isinstance(data_val, datetime):
                    data = data_val
                elif isinstance(data_val, float):
                    data = _xl_date(data_val)
                elif isinstance(data_val, str) and data_val:
                    try:
                        from datetime import datetime as dt2
                        data = dt2.fromisoformat(data_val.split(' ')[0])
                    except: data = None
                else:
                    data = None

                nf_val = row[col_nf]
                nf = str(int(nf_val)) if isinstance(nf_val, float) else str(nf_val).split('.')[0]

                prod_norm = normalizar_produto(prod)
                notas.append({
                    'nf': nf, 'data': data,
                    'produto_raw': prod,
                    'produto': prod_norm,
                    'quantidade': qtd,
                })
            except: continue

    except Exception as e:
        return None, str(e)

    if not notas:
        return None, "Nenhuma nota de combustível encontrada. Verifique se o arquivo contém linhas de 'Entrada' com unidade 'L'."

    totais = {}
    for n in notas:
        p = n['produto']
        totais[p] = totais.get(p, 0) + n['quantidade']

    return {'notas': notas, 'totais': totais}, None


def ler_estoque_dac(arquivo_bytes, filename):
    """
    Lê o Estoque Físico do DAC (PDF Resumo DAC).
    Retorna: lista de tanques com est_aber, est_fech, recebido, faltas_sobras.
    """
    try:
        reader = pypdf.PdfReader(io.BytesIO(arquivo_bytes))
        texto  = "".join(p.extract_text() or "" for p in reader.pages)
    except Exception as e:
        return None, str(e)

    if len(texto.strip()) < 50:
        return None, "PDF escaneado ou sem texto — não é possível extrair dados."

    # Competência
    competencia = ""
    m = re.search(r'(\d{2}/\d{2}/\d{4})\s+at[ée]\s+(\d{2}/\d{2}/\d{4})', texto)
    if m:
        dt = m.group(2)
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try: competencia = f"{meses[int(dt[3:5])-1]}/{dt[6:]}"
        except: competencia = dt

    # Nome da empresa
    empresa = ""
    m_emp = re.search(r"Resumo\s+DAC\s*\n\s*(.+?)\s+Emiss", texto, re.I)
    if m_emp:
        empresa = m_emp.group(1).strip()
    else:
        m2 = re.search(r"([\w\s]+(?:LTDA|EIRELI|EPP|ME))", texto, re.I)
        if m2: empresa = m2.group(1).strip()

    # Extrair Estoque Físico
    num = r'-?[\d\.]+,\d+'
    linhas = texto.splitlines()
    em_estoque = False
    tanques = []

    for l in linhas:
        l = l.strip()
        if re.search(r'Estoque\s+F[íi]sico', l, re.I): em_estoque = True; continue
        if re.search(r'Resumo\s+DAC\s*-|P[áa]gina\s+\d', l, re.I): em_estoque = False; continue
        if not em_estoque: continue

        # Formato 3 valores: "001 0,000GASOLINA COMUM 3.173,787 10.000,000 10.973,295"
        #   → est_fech | recebido | faltas_sobras
        # Formato 2 valores: "003 0,000ETANOL COMUM 1.662,096 4.975,156"
        #   → est_fech | faltas_sobras (recebido = 0, tanque não recebeu no mês)
        pat3 = re.compile(
            r'^(\d{3})\s+(-?[\d\.]+,\d+)\s*([A-ZÀ-Ú][A-ZÀ-Ú0-9 \-]+?)\s+'
            r'(' + num + r')\s+(' + num + r')\s+(-?' + num + r')\s*$'
        )
        pat2 = re.compile(
            r'^(\d{3})\s+(-?[\d\.]+,\d+)\s*([A-ZÀ-Ú][A-ZÀ-Ú0-9 \-]+?)\s+'
            r'(' + num + r')\s+(-?' + num + r')\s*$'
        )
        m3 = pat3.match(l)
        m2 = pat2.match(l) if not m3 else None

        if m3:
            tanques.append({
                'tanque':        m3.group(1),
                'produto_raw':   m3.group(3).strip(),
                'produto':       normalizar_produto(m3.group(3)),
                'est_aber':      _fl(m3.group(2)),
                'est_fech':      _fl(m3.group(4)),
                'recebido':      _fl(m3.group(5)),
                'faltas_sobras': _fl(m3.group(6)),
            })
        elif m2:
            # Apenas 2 valores numéricos após o produto: est_fech e faltas_sobras
            # Recebido = 0 pois não houve recebimento neste tanque no período
            tanques.append({
                'tanque':        m2.group(1),
                'produto_raw':   m2.group(3).strip(),
                'produto':       normalizar_produto(m2.group(3)),
                'est_aber':      _fl(m2.group(2)),
                'est_fech':      _fl(m2.group(4)),
                'recebido':      0.0,
                'faltas_sobras': _fl(m2.group(5)),
            })

    if not tanques:
        return None, "Seção 'Estoque Físico' não encontrada no DAC."

    # Consolidar por produto — soma recebido, mas separa perda e ganho por tanque
    # para não compensar perda de um tanque com ganho de outro
    consolidado = {}
    for t in tanques:
        p = t['produto']
        if p not in consolidado:
            consolidado[p] = {
                'produto': p, 'produto_raw': t['produto_raw'],
                'tanques': [], 'est_aber': 0, 'est_fech': 0,
                'recebido': 0, 'perda': 0, 'ganho': 0,
            }
        fs = t['faltas_sobras'] or 0
        consolidado[p]['tanques'].append(t['tanque'])
        consolidado[p]['est_aber']  += t['est_aber']  or 0
        consolidado[p]['est_fech']  += t['est_fech']  or 0
        consolidado[p]['recebido']  += t['recebido']  or 0
        # Separar: cada tanque contribui individualmente para perda ou ganho
        if fs < 0:
            consolidado[p]['perda'] += abs(fs)
        elif fs > 0:
            consolidado[p]['ganho'] += fs

    return {
        'competencia': competencia,
        'empresa': empresa,
        'tanques_bruto': tanques,
        'por_produto': consolidado,
    }, None


# ── Confronto Notas × DAC ─────────────────────────────────────────────────────
def confrontar_recebimento(notas_data, dac_data):
    """
    Cruza totais das notas fiscais com o Estoque Físico do DAC.
    Retorna lista de linhas de confronto por produto.
    """
    todos_produtos = set(list(notas_data['totais'].keys()) +
                         list(dac_data['por_produto'].keys()))

    linhas = []
    for p in sorted(todos_produtos):
        qtd_notas = notas_data['totais'].get(p, 0)
        dac_item  = dac_data['por_produto'].get(p, {})
        rec_dac   = dac_item.get('recebido', 0) or 0
        perda     = dac_item.get('perda', 0) or 0
        ganho     = dac_item.get('ganho', 0) or 0
        est_aber  = dac_item.get('est_aber', None)
        est_fech  = dac_item.get('est_fech', None)

        # Diferença notas × DAC
        dif = round(qtd_notas - rec_dac, 3) if rec_dac or qtd_notas else None

        # Status
        if qtd_notas == 0 and rec_dac == 0:
            status = '—'
        elif dif is None:
            status = '⚠️ Sem dado DAC'
        elif abs(dif) < 0.01:
            status = '✅ OK'
        elif dif > 0:
            status = '⚠️ Nota > DAC'
        else:
            status = '⚠️ DAC > Nota'

        linhas.append({
            'produto':   p,
            'qtd_notas': qtd_notas,
            'rec_dac':   rec_dac,
            'diferenca': dif,
            'est_aber':  est_aber,
            'est_fech':  est_fech,
            'perda':     perda,
            'ganho':     ganho,
            'status':    status,
            'tanques':   dac_item.get('tanques', []),
        })

    return linhas