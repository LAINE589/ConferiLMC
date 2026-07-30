import os
import io, re, unicodedata
from datetime import datetime, timedelta
from functools import wraps

from flask import (Flask, render_template, request, redirect,
                   url_for, session, send_file, flash, abort)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, login_user, logout_user,
                         login_required as fl_login_required,
                         current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, Usuario, Posto, Contabilidade, Relatorio
from models import PERFIL_ADMIN, PERFIL_CONTABILIDADE, PERFIL_POSTO

import openpyxl
try:
    from dac_reader import ler_dac, confrontar_dac_sped
    DAC_DISPONIVEL = True
except Exception:
    DAC_DISPONIVEL = False
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "LMC-SPED-2026-K9x#mQpZ")
app.permanent_session_lifetime = timedelta(hours=8)

# ── Banco de dados ────────────────────────────────────────────────────────────
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    f"sqlite:///{os.path.join(BASE_DIR, 'lmc.db')}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db.init_app(app)

# ── Flask-Login ───────────────────────────────────────────────────────────────
login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "Faça login para continuar."
login_manager.login_message_category = "warning"

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# ── Pasta de relatórios por posto ─────────────────────────────────────────────
RELATORIOS_DIR = os.path.join(BASE_DIR, "relatorios")
os.makedirs(RELATORIOS_DIR, exist_ok=True)

def _nome_posto(info_atu):
    """Sanitiza o nome da empresa para uso em nome de arquivo."""
    razao = info_atu.get("info", {}).get("razao", "") or "RELATORIO_LMC"
    razao = unicodedata.normalize("NFKD", razao)
    razao = "".join(c for c in razao if not unicodedata.combining(c))
    razao = re.sub(r"[^A-Za-z0-9 ]", "", razao).strip()
    razao = re.sub(r"\s+", "_", razao)
    return razao[:40] or "RELATORIO_LMC"

# ── Inicializar banco e criar admin padrão ────────────────────────────────────
def init_db():
    with app.app_context():
        db.create_all()
        # Criar admin principal se não existir
        if not Usuario.query.filter_by(email="lainerose1994@gmail.com").first():
            admin = Usuario(
                email="lainerose1994@gmail.com",
                nome="Laíne Rose",
                perfil=PERFIL_ADMIN,
                ativo=True,
            )
            admin.set_senha("Cld@Admin2026!")
            db.session.add(admin)

        # Garantir contabilidade padrão
        cont = Contabilidade.query.filter_by(cnpj="00000000000000").first()
        if not cont:
            cont = Contabilidade(
                nome="Cleodon Contabilidade",
                cnpj="00000000000000",
                email="fiscal@cleodoncontabilidade.com.br",
            )
            db.session.add(cont)
            db.session.flush()

        # Criar usuários de contabilidade se não existirem
        for email, senha, nome in [
            ("fiscal@cleodoncontabilidade.com.br", "Cld@123", "Fiscal"),
            ("lucroreal@cleodoncontabilidade.com.br", "Cld@123", "Lucro Real"),
        ]:
            if not Usuario.query.filter_by(email=email).first():
                u = Usuario(email=email, nome=nome,
                            perfil=PERFIL_CONTABILIDADE,
                            contabilidade_id=cont.id)
                u.set_senha(senha)
                db.session.add(u)

        db.session.commit()

init_db()

# ── Decoradores de acesso ─────────────────────────────────────────────────────
def requer_perfil(*perfis):
    """Decorador que exige um ou mais perfis específicos."""
    def decorator(f):
        @wraps(f)
        @fl_login_required
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated or current_user.perfil not in perfis:
                flash("Acesso não autorizado.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─────────────────────────────────────────────────────────────────────────────
# ROTAS DE AUTENTICAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    erro = None
    if request.method == "POST":
        email = request.form.get("usuario", "").strip().lower()
        senha = request.form.get("senha", "").strip()
        user  = Usuario.query.filter_by(email=email, ativo=True).first()
        if user and user.check_senha(senha):
            login_user(user, remember=True)
            user.ultimo_acesso = datetime.utcnow()
            db.session.commit()
            return redirect(url_for("dashboard"))
        erro = "Usuário ou senha incorretos. Verifique e tente novamente."
    return render_template("login.html", erro=erro)


@app.route("/logout")
@fl_login_required
def logout():
    logout_user()
    flash("Você saiu do sistema.", "success")
    return redirect(url_for("login"))


@app.route("/dashboard")
@fl_login_required
def dashboard():
    """Redireciona para o painel correto conforme o perfil."""
    if current_user.is_admin:
        return redirect(url_for("admin_dashboard"))
    if current_user.is_contabilidade:
        return redirect(url_for("index"))
    if current_user.is_posto:
        # Verificar licença
        posto = current_user.posto
        if not posto or not posto.licenca_ativa:
            flash("Sua licença está inativa. Entre em contato com a Cleodon Contabilidade.", "danger")
            return redirect(url_for("login"))
        return redirect(url_for("posto_dashboard"))
    return redirect(url_for("login"))


@app.route("/sistema")
@fl_login_required
def index():
    """Painel de conferência — acessível por Admin e Contabilidade."""
    if current_user.is_posto:
        return redirect(url_for("posto_dashboard"))
    # Admin e Contabilidade acessam a interface de conferência normalmente
    return render_template("index.html", nome=current_user.nome, usuario=current_user)


@app.route("/processar", methods=["POST"])
@fl_login_required
def processar():
    arq_ant = request.files.get("ant")
    arq_atu = request.files.get("atu")

    # SPED da competência anterior agora é opcional.
    # Se ausente, o sistema confronta apenas o mês atual com o DAC,
    # mantendo consistência diária, negativos, versão/capacidade e ANP.
    tem_ant = bool(arq_ant and arq_ant.filename)

    if not arq_atu or not arq_atu.filename:
        flash("Selecione ao menos o arquivo SPED da competência atual.", "danger")
        return redirect(url_for("index"))

    try:
        bytes_atu = arq_atu.read()
        d_atu = ler_sped_bytes(bytes_atu)

        if tem_ant:
            bytes_ant = arq_ant.read()
            d_ant   = ler_sped_bytes(bytes_ant)
            neg_abr = verificar_negativos_bytes(bytes_ant)
        else:
            # Estrutura vazia equivalente — confronto_mensal já trata isso
            # retornando listas vazias quando não há dados do mês anterior.
            d_ant   = {"info": {}, "tanques": {}, "bicos": {}}
            neg_abr = {"tanques": [], "bicos": []}
            flash("SPED da competência anterior não enviado — gerando apenas a "
                  "conferência da competência atual (sem confronto entre meses).",
                  "warning")

        conf_m  = confronto_mensal(d_ant, d_atu)
        d_mai   = confronto_diario(d_atu)
        neg_mai = verificar_negativos_bytes(bytes_atu)
        vc_mai  = verificar_versao_capacidade(d_atu)

        # DAC opcional
        arq_dac = request.files.get("dac")
        conf_dac = None
        erro_dac = None
        if arq_dac and arq_dac.filename:
            try:
                bytes_dac = arq_dac.read()
                dac_dados = ler_dac(bytes_dac, arq_dac.filename)
                conf_dac  = confrontar_dac_sped(dac_dados, d_atu)
            except Exception as e:
                erro_dac = str(e)

        cad_atu = verificar_cadastro_lmc(bytes_atu)

        wb  = openpyxl.Workbook()
        ws1 = wb.active; ws1.title = "Resumo"
        aba_resumo(ws1, conf_m, d_mai, d_ant, d_atu, neg_abr, neg_mai, vc_mai, cad_atu)
        ws2 = wb.create_sheet("Confronto Meses")
        aba_mensal(ws2, conf_m, d_ant, d_atu)
        ws3 = wb.create_sheet("Comparativo Diário")
        aba_diario(ws3, d_mai)
        if conf_dac:
            ws4 = wb.create_sheet("DAC × SPED")
            aba_dac(ws4, conf_dac, d_atu)
            ws5 = wb.create_sheet("DAC do SPED")
            aba_dac_sped(ws5, d_atu, d_atu)
        else:
            # Sem DAC real: gerar DAC de acompanhamento a partir do SPED
            ws4 = wb.create_sheet("DAC do SPED")
            aba_dac_sped(ws4, d_atu, d_atu)
        if erro_dac:
            flash(f"Aviso DAC: {erro_dac}", "warning")

        ws_rel = wb.create_sheet("Relatório ao Cliente")
        aba_relatorio_cliente(ws_rel, conf_m, d_mai, neg_abr, neg_mai, vc_mai,
                              conf_dac, d_ant, d_atu, cad_atu)

        ws_cad = wb.create_sheet("Informação das Bombas", 1)
        aba_cadastro_lmc(ws_cad, cad_atu, d_atu)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        nome = f"Relatorio_LMC_{_nome_posto(d_atu)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=nome)

    except Exception as e:
        flash(f"Erro ao processar os arquivos: {str(e)}", "danger")
        return redirect(url_for("index"))


# ─────────────────────────────────────────────────────────────────────────────from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

C_AZUL_ESC = "A81550"; C_AZUL_MED = "C41A60"
C_VERDE_BG = "C6EFCE"; C_VERDE_FG = "375623"
C_VERM_BG  = "FFC7CE"; C_VERM_FG  = "9C0006"
C_AMAR_BG  = "FFEB9C"; C_AMAR_FG  = "7D6608"
C_CINZA    = "D9D9D9"
NF = "#,##0.000"
VERSAO_OBRIGATORIA = "020"

def _sk(x):
    return int(x) if str(x).isdigit() else x

def _nid(raw):
    try:
        return str(int(raw.strip()))
    except:
        return raw.strip()

def _brd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _fl(v):
    return float(v.strip().replace(",", ".")) if v.strip() else 0.0

def _dt(s):
    s = s.strip()
    return datetime.strptime(s, "%d%m%Y").date() if len(s) == 8 else None

def _st(dif):
    if dif is None: return "⚠️ AUSENTE"
    return "✅ OK" if dif == 0 else "❌ DIVERGÊNCIA"

def _row_bg(status):
    return {"✅ OK": C_VERDE_BG, "❌ DIVERGÊNCIA": C_VERM_BG, "⚠️ AUSENTE": C_AMAR_BG}.get(status)

def _label(mapa, id_raw, tipo):
    """Retorna 'Tanque 1', 'Bico 3' etc. usando o mapa ordinal."""
    return f"{tipo} {mapa.get(id_raw, id_raw)}"

# ── LEITURA ────────────────────────────────────────────────────────────────────
def ler_sped(caminho):
    with open(caminho, encoding="latin-1", errors="replace") as f:
        text = f.read()
    tanques={}; bicos={}; info={}
    data_atual=None; vals_1300={}

    for linha in text.splitlines():
        c=linha.strip().split("|")
        if len(c)<2: continue
        tp=c[1]

        if tp=="0000":
            info={"versao":c[2].strip() if len(c)>2 else "",
                  "razao": c[6].strip() if len(c)>6 else "",
                  "cnpj":  c[7].strip() if len(c)>7 else "",
                  "dt_ini":c[4].strip() if len(c)>4 else "",
                  "dt_fin":c[5].strip() if len(c)>5 else ""}

        elif tp=="1300":
            data_atual=_dt(c[3]) if len(c)>3 else None
            if not data_atual: continue
            vals_1300={
                "data":data_atual,
                "est_abert":_fl(c[4]),"entrada":_fl(c[5]),"saida":_fl(c[7]),
                "evap":_fl(c[9]) if len(c)>9 else 0.0,
                "ajuste":_fl(c[10]) if len(c)>10 else 0.0,
                "est_fech":_fl(c[11]) if len(c)>11 else 0.0,
            }

        elif tp=="1310":
            if not vals_1300: continue
            t=_nid(c[2])
            tem_campos_proprios = len(c) > 10 and c[3].strip() != ""
            if tem_campos_proprios:
                try:
                    est_abert_1310 = _fl(c[3])
                    entrada_1310   = _fl(c[4]) if len(c)>4 else 0.0
                    saida_1310     = _fl(c[6]) if len(c)>6 else 0.0
                    evap_1310      = _fl(c[8]) if len(c)>8 else 0.0
                    ajuste_1310    = _fl(c[9]) if len(c)>9 else 0.0
                    est_fech_1310  = _fl(c[10]) if len(c)>10 else 0.0
                    cap            = _fl(c[11]) if len(c)>11 and c[11].strip() else None
                except Exception:
                    tem_campos_proprios = False

            if tem_campos_proprios:
                key=(t,vals_1300["data"])
                tanques[key]={
                    "tanque":t,"data":vals_1300["data"],
                    "est_abert":est_abert_1310,"entrada":entrada_1310,
                    "saida":saida_1310,"evap":evap_1310,
                    "ajuste":ajuste_1310,"est_fech":est_fech_1310,
                    "capacidade":cap,
                }
            else:
                cap=_fl(c[11]) if len(c)>11 and c[11].strip() else None
                key=(t,vals_1300["data"])
                tanques[key]={
                    "tanque":t,"data":vals_1300["data"],
                    "est_abert":vals_1300["est_abert"],"entrada":vals_1300["entrada"],
                    "saida":vals_1300["saida"],"evap":vals_1300["evap"],
                    "ajuste":vals_1300["ajuste"],"est_fech":vals_1300["est_fech"],
                    "capacidade":cap,
                }

        elif tp=="1320":
            b=_nid(c[2])
            if not data_atual: continue
            bicos[(b,data_atual)]={
                "bico":b,"data":data_atual,
                "enc_abert":_fl(c[9]) if len(c)>9 else 0.0,
                "enc_fech": _fl(c[8]) if len(c)>8 else 0.0,
            }

    return {"info":info,"tanques":tanques,"bicos":bicos}


def confronto_mensal(d_ant, d_atu):
    res = {"tanques": [], "bicos": []}
    dt_ant = sorted(set(d for (_,d) in d_ant["tanques"]))
    dt_atu = sorted(set(d for (_,d) in d_atu["tanques"]))
    if not dt_ant or not dt_atu: return res
    ult=dt_ant[-1]; pri=dt_atu[0]

    tanq_pri={}; tanq_ult={}
    for (t,d) in d_atu["tanques"]:
        if t not in tanq_pri or d<tanq_pri[t]: tanq_pri[t]=d
        if t not in tanq_ult or d>tanq_ult[t]: tanq_ult[t]=d

    for t in sorted(set([t for (t,d) in d_ant["tanques"] if d==ult]+
                        list(tanq_pri.keys())), key=_sk):
        fa   = d_ant["tanques"].get((t,ult),{})
        fech = fa.get("est_fech")
        pri_t=tanq_pri.get(t); ult_t=tanq_ult.get(t)
        aa   = d_atu["tanques"].get((t,pri_t),{}) if pri_t else {}
        fu   = d_atu["tanques"].get((t,ult_t),{}) if ult_t else {}
        aber     = aa.get("est_abert")
        fech_atu = fu.get("est_fech")
        dif = round(aber-fech,3) if fech is not None and aber is not None else None

        # Cálculo ANP
        dias_t = sorted(d for (tt,d) in d_atu["tanques"] if tt==t)
        est_ini    = d_atu["tanques"][(t,dias_t[0])]["est_abert"]  if dias_t else None
        est_fin_sp = d_atu["tanques"][(t,dias_t[-1])]["est_fech"]  if dias_t else None
        total_rec  = round(sum(d_atu["tanques"][(t,d)]["entrada"] for d in dias_t),3)
        total_sai  = round(sum(d_atu["tanques"][(t,d)]["saida"]   for d in dias_t),3)
        if est_ini is not None and est_fin_sp is not None:
            saldo_calc    = round(est_ini+total_rec-total_sai,3)
            diferenca_anp = round(saldo_calc-est_fin_sp,3)
            limite_anp    = round(total_rec*0.006,3)
            pct_anp       = round(abs(diferenca_anp)/total_rec*100,4) if total_rec>0 else 0
            if abs(diferenca_anp)<=limite_anp:
                status_anp="✅ DENTRO DO LIMITE"
            elif diferenca_anp>0:
                status_anp="⚠️ SOBRA ACIMA 0,6%"
            else:
                status_anp="❌ FALTA ACIMA 0,6%"
        else:
            diferenca_anp=limite_anp=pct_anp=None; status_anp="⚠️ AUSENTE"

        res["tanques"].append({
            "id":t,"dt_fech":ult,"fech":fech,"dt_aber":pri_t,"aber":aber,
            "dif":dif,"status":_st(dif),"dt_fech_atu":ult_t,"fech_atu":fech_atu,
            "total_rec":total_rec,"total_sai":total_sai,
            "diferenca_anp":diferenca_anp,"limite_anp":limite_anp,
            "pct_anp":pct_anp,"status_anp":status_anp,
        })

    bico_pri={}; bico_ult={}
    for (b,d) in d_atu["bicos"]:
        if b not in bico_pri or d<bico_pri[b]: bico_pri[b]=d
        if b not in bico_ult or d>bico_ult[b]: bico_ult[b]=d

    db_ant=sorted(set(d for (_,d) in d_ant["bicos"]))
    ult_b=db_ant[-1] if db_ant else ult

    for b in sorted(set([b for (b,d) in d_ant["bicos"] if d==ult_b]+
                        list(bico_pri.keys())), key=_sk):
        fa  = d_ant["bicos"].get((b,ult_b),{})
        fech= fa.get("enc_fech")
        pri_b=bico_pri.get(b); ult_b2=bico_ult.get(b)
        aa  = d_atu["bicos"].get((b,pri_b),{}) if pri_b else {}
        fu  = d_atu["bicos"].get((b,ult_b2),{}) if ult_b2 else {}
        aber    = aa.get("enc_abert")
        fech_atu= fu.get("enc_fech")
        dif = round(aber-fech,3) if fech is not None and aber is not None else None
        res["bicos"].append({
            "id":b,"dt_fech":ult_b,"fech":fech,"dt_aber":pri_b,"aber":aber,
            "dif":dif,"status":_st(dif),"dt_fech_atu":ult_b2,"fech_atu":fech_atu,
        })
    return res

def confronto_diario(dados):
    res_t = []; res_b = []
    for tanque in sorted(set(t for (t,_) in dados["tanques"]), key=lambda x: int(x) if x.isdigit() else x):
        dias = sorted(d for (t,d) in dados["tanques"] if t==tanque)
        for i in range(len(dias)-1):
            d1,d2 = dias[i],dias[i+1]
            r1=dados["tanques"][(tanque,d1)]; r2=dados["tanques"][(tanque,d2)]
            fech=r1["est_fech"]; aber=r2["est_abert"]; dif=round(aber-fech,3)
            res_t.append({"tanque":tanque,"dia_fech":d1,"fech":fech,"dia_aber":d2,"aber":aber,"dif":dif,"status":_st(dif),})
    for bico in sorted(set(b for (b,_) in dados["bicos"]), key=lambda x: int(x) if x.isdigit() else x):
        dias = sorted(d for (b,d) in dados["bicos"] if b==bico)
        for i in range(len(dias)-1):
            d1,d2 = dias[i],dias[i+1]
            r1=dados["bicos"][(bico,d1)]; r2=dados["bicos"][(bico,d2)]
            fech=r1["enc_fech"]; aber=r2["enc_abert"]; dif=round(aber-fech,3)
            res_b.append({"bico":bico,"dia_fech":d1,"fech":fech,"dia_aber":d2,"aber":aber,"dif":dif,"status":_st(dif),})
    return {"tanques": res_t, "bicos": res_b}

# ── NEGATIVOS ──────────────────────────────────────────────────────────────────
CAMPOS_1300 = {5:"Est. Abertura",6:"Entrada",7:"Est. Aber. Pós Entrada",8:"Saída",
               9:"Est. Fech. Pré Ajuste",10:"Evaporação",11:"Ajuste",12:"Est. Fechamento Final"}
CAMPOS_1320 = {8:"Enc. Fechamento",9:"Enc. Abertura",10:"Volume Vendido",11:"Diferença Encerrante"}

def verificar_negativos(caminho):
    neg_t=[]; neg_b=[]; data_atual=None
    with open(caminho, encoding="latin-1", errors="replace") as f:
        for n, linha in enumerate(f, 1):
            c = linha.strip().split("|")
            if len(c)<2: continue
            tp=c[1]
            if tp=="1300":
                data_atual=_dt(c[3]) if len(c)>3 else None
                tanque=_nid(c[2])
                for idx,nome in CAMPOS_1300.items():
                    if idx>=len(c): continue
                    try:
                        v=_fl(c[idx])
                        if v<0: neg_t.append({"tanque":tanque,"data":data_atual,"campo":nome,"valor":v,"linha":n})
                    except: pass
            elif tp=="1320":
                bico=_nid(c[2])
                for idx,nome in CAMPOS_1320.items():
                    if idx>=len(c): continue
                    try:
                        v=_fl(c[idx])
                        if v<0: neg_b.append({"bico":bico,"data":data_atual,"campo":nome,"valor":v,"linha":n})
                    except: pass
    return {"tanques": neg_t, "bicos": neg_b}

# ── VERSÃO E CAPACIDADE ────────────────────────────────────────────────────────
def verificar_versao_capacidade(dados):
    info=dados["info"]; versao=info.get("versao",""); dt_ini=info.get("dt_ini","")
    periodo=f"{dt_ini[2:4]}/{dt_ini[4:]}" if len(dt_ini)==8 else dt_ini
    ano=int(dt_ini[4:]) if len(dt_ini)==8 else 0
    cap_obrig = ano >= 2026
    tanques_ids=sorted(set(t for (t,_) in dados["tanques"]), key=lambda x: int(x) if x.isdigit() else x)
    cap_tanques=[]
    for t in tanques_ids:
        caps=set()
        for d in sorted(d for (tt,d) in dados["tanques"] if tt==t):
            c=dados["tanques"].get((t,d),{}).get("capacidade")
            if c is not None: caps.add(c)
        if not caps:
            st="❌ AUSENTE" if cap_obrig else "⚠️ NÃO DECLARADA"; obs="Capacidade não informada"
        elif len(caps)>1:
            st="⚠️ INCONSISTENTE"; obs=f"Valores distintos: {sorted(caps)}"
        elif list(caps)[0]<=0:
            st="❌ INVÁLIDA"; obs=f"Valor zero/negativo: {list(caps)[0]}"
        else:
            st="✅ OK"; obs=f"{list(caps)[0]:,.0f} L"
        cap_tanques.append({"tanque":t,"caps":sorted(caps),"status":st,"obs":obs})
    return {"versao":versao,"versao_ok":versao==VERSAO_OBRIGATORIA,
            "periodo":periodo,"cap_obrig":cap_obrig,"tanques":cap_tanques}

# ── HELPERS EXCEL ──────────────────────────────────────────────────────────────
def _brd_cel(ws,r,c,val,bg=None,fg="000000",bold=False,sz=10,align="center",wrap=False,fmt=None):
    cel=ws.cell(row=r,column=c,value=val)
    cel.font=Font(name="Arial",size=sz,bold=bold,color=fg)
    cel.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    cel.border=_brd()
    if bg: cel.fill=PatternFill("solid",start_color=bg)
    if fmt: cel.number_format=fmt
    return cel

def _ch(ws,r,c,val,bg=C_AZUL_ESC,fg="FFFFFF",sz=10):
    cel=ws.cell(row=r,column=c,value=val)
    cel.font=Font(name="Arial",bold=True,color=fg,size=sz)
    cel.fill=PatternFill("solid",start_color=bg)
    cel.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cel.border=_brd(); return cel

def _dc(ws,r,c,val,fmt=None,bg=None):
    return _brd_cel(ws,r,c,val,bg=bg,fmt=fmt)

def _sc(ws,r,c,status):
    mapa={"✅ OK":(C_VERDE_BG,C_VERDE_FG),"❌ DIVERGÊNCIA":(C_VERM_BG,C_VERM_FG),
          "⚠️ AUSENTE":(C_AMAR_BG,C_AMAR_FG),"❌ AUSENTE":(C_VERM_BG,C_VERM_FG),
          "⚠️ INCONSISTENTE":(C_AMAR_BG,C_AMAR_FG),"❌ INVÁLIDA":(C_VERM_BG,C_VERM_FG)}
    bg,fg=mapa.get(status,(C_CINZA,"000000"))
    cel=ws.cell(row=r,column=c,value=status)
    cel.font=Font(name="Arial",bold=True,size=10,color=fg)
    cel.fill=PatternFill("solid",start_color=bg)
    cel.alignment=Alignment(horizontal="center",vertical="center")
    cel.border=_brd(); return cel

def _titulo(ws,r,texto,n,sz=12):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n)
    c=ws.cell(row=r,column=1,value=texto)
    c.font=Font(name="Arial",bold=True,size=sz,color="FFFFFF")
    c.fill=PatternFill("solid",start_color=C_AZUL_ESC)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[r].height=26; return c

def _subtit(ws,r,texto,n):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=n)
    c=ws.cell(row=r,column=1,value=texto)
    c.font=Font(name="Arial",bold=True,size=10,color="FFFFFF")
    c.fill=PatternFill("solid",start_color=C_AZUL_MED)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[r].height=20; return c

# ── ABA RESUMO ─────────────────────────────────────────────────────────────────
def _inserir_logo(ws, ancora="A1", altura=36):
    """Insere a logo da Cleodon Contabilidade na célula indicada."""
    try:
        from openpyxl.drawing.image import Image as XlImage
        logo_path = os.path.join(os.path.dirname(__file__), "static", "logo_excel.png")
        if os.path.exists(logo_path):
            logo = XlImage(logo_path)
            logo.width  = int(altura * 0.97)
            logo.height = altura
            logo.anchor = ancora
            ws.add_image(logo)
    except Exception:
        pass

def aba_resumo(ws, conf_m, d_mai, info_ant, info_atu, neg_abr, neg_mai, vc_mai, cad_atu=None):
    ws.sheet_view.showGridLines=False; N=7
    _titulo(ws,1,"CONFERÊNCIA LMC – LIVRO DE MOVIMENTAÇÃO DE COMBUSTÍVEIS",N,sz=13)
    ws.row_dimensions[1].height=40
    # Logo no canto direito do cabeçalho
    _inserir_logo(ws, "G1")
    ia=info_ant["info"]; iu=info_atu["info"]
    razao_emp = ia.get('razao','') or iu.get('razao','')
    cnpj_emp  = ia.get('cnpj','')  or iu.get('cnpj','')
    comp_ant_txt = (f"{ia.get('dt_ini','')} a {ia.get('dt_fin','')}"
                    if ia.get('dt_ini') else "Não enviado (apenas competência atual)")
    for i,(a,e) in enumerate([(f"Empresa: {razao_emp}", f"CNPJ: {cnpj_emp}"),
        (f"Competência anterior: {comp_ant_txt}", f"Competência atual: {iu.get('dt_ini','')} a {iu.get('dt_fin','')}"),
        (f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", "")], start=2):
        ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
        c1=ws.cell(row=i,column=1,value=a); c1.font=Font(name="Arial",bold=(i==2),size=10)
        c1.alignment=Alignment(horizontal="left",vertical="center")
        ws.merge_cells(start_row=i,start_column=5,end_row=i,end_column=N)
        c2=ws.cell(row=i,column=5,value=e); c2.font=Font(name="Arial",size=10)
        c2.alignment=Alignment(horizontal="right",vertical="center")
        ws.row_dimensions[i].height=16

    def bloco_ok(r, texto):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=N)
        c=ws.cell(row=r,column=1,value=f"✅  {texto}")
        c.font=Font(name="Arial",bold=True,size=10,color=C_VERDE_FG)
        c.fill=PatternFill("solid",start_color=C_VERDE_BG)
        c.alignment=Alignment(horizontal="left",vertical="center"); c.border=_brd()
        ws.row_dimensions[r].height=18; return r+1

    def bloco_err(r, texto):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=N)
        c=ws.cell(row=r,column=1,value=f"❌  {texto}")
        c.font=Font(name="Arial",bold=True,size=10,color=C_VERM_FG)
        c.fill=PatternFill("solid",start_color=C_VERM_BG)
        c.alignment=Alignment(horizontal="left",vertical="center"); c.border=_brd()
        ws.row_dimensions[r].height=18; return r+1

    def detalhe(r, texto, cor_bg=C_VERM_BG, cor_fg=C_VERM_FG):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=N)
        c=ws.cell(row=r,column=1,value=f"     ▸  {texto}")
        c.font=Font(name="Arial",size=10,color=cor_fg)
        c.fill=PatternFill("solid",start_color=cor_bg)
        c.alignment=Alignment(horizontal="left",vertical="center"); c.border=_brd()
        ws.row_dimensions[r].height=16; return r+1

    row=6

    # 1. CONFRONTO MENSAL
    _subtit(ws,row,"1.  CONFRONTO ENTRE MESES  –  Fechamento Anterior × Abertura Atual",N); row+=1
    sem_anterior = not conf_m["tanques"] and not conf_m["bicos"]
    if sem_anterior:
        row=bloco_ok(row, "SPED da competência anterior não enviado — confronto entre meses não realizado. "
                          "Conferência abaixo considera apenas a competência atual.")
    else:
        for i,h in enumerate(["Tipo","ID","Data Fech.","Valor Fech.","Data Aber.","Valor Aber.","Status"],1):
            _ch(ws,row,i,h,bg=C_CINZA,fg=C_AZUL_ESC)
        ws.row_dimensions[row].height=22; row+=1
        for tipo, lista in [("Tanque", conf_m["tanques"]), ("Bico", conf_m["bicos"])]:
            for x in lista:
                bg=_row_bg(x["status"])
                _dc(ws,row,1,tipo,bg=bg)
                _dc(ws,row,2,f"{tipo} {x['id']}",bg=bg)
                _dc(ws,row,3,x["dt_fech"].strftime("%d/%m/%Y") if x.get("dt_fech") else "",bg=bg)
                _dc(ws,row,4,x["fech"],NF,bg=bg)
                _dc(ws,row,5,x["dt_aber"].strftime("%d/%m/%Y") if x.get("dt_aber") else "",bg=bg)
                _dc(ws,row,6,x["aber"],NF,bg=bg); _sc(ws,row,7,x["status"])
                ws.row_dimensions[row].height=15; row+=1
    row+=1

    # 2. CONSISTÊNCIA DIÁRIA
    _subtit(ws,row,"2.  CONSISTÊNCIA DIÁRIA – COMPETÊNCIA ATUAL  (Fechamento dia N = Abertura dia N+1)",N); row+=1
    divs_t=[x for x in d_mai["tanques"] if x["status"]=="❌ DIVERGÊNCIA"]
    divs_b=[x for x in d_mai["bicos"]   if x["status"]=="❌ DIVERGÊNCIA"]

    if not divs_t:
        row=bloco_ok(row, f"Tanques: {len(d_mai['tanques'])} transições verificadas — nenhuma divergência encontrada")
    else:
        row=bloco_err(row, f"Tanques: {len(divs_t)} divergência(s) de {len(d_mai['tanques'])} transições")
        for x in divs_t:
            row=detalhe(row, f"Tanque {x['tanque']}  |  Fech. {x['dia_fech'].strftime('%d/%m/%Y')}: {x['fech']:,.3f}  →  Aber. {x['dia_aber'].strftime('%d/%m/%Y')}: {x['aber']:,.3f}  |  Dif.: {x['dif']:,.3f} L")

    if not divs_b:
        row=bloco_ok(row, f"Bicos: {len(d_mai['bicos'])} transições verificadas — nenhuma divergência encontrada")
    else:
        row=bloco_err(row, f"Bicos: {len(divs_b)} divergência(s) de {len(d_mai['bicos'])} transições")
        for x in divs_b:
            row=detalhe(row, f"Bico {x['bico']}  |  Fech. {x['dia_fech'].strftime('%d/%m/%Y')}: {x['fech']:,.3f}  →  Aber. {x['dia_aber'].strftime('%d/%m/%Y')}: {x['aber']:,.3f}  |  Dif.: {x['dif']:,.3f}")
    row+=1

    # 3. VALORES NEGATIVOS
    _subtit(ws,row,"3.  VALORES NEGATIVOS NOS REGISTROS 1310 / 1320  (Ambas as competências)",N); row+=1
    todos_neg_t = neg_abr["tanques"] + neg_mai["tanques"]
    todos_neg_b = neg_abr["bicos"]   + neg_mai["bicos"]

    if not todos_neg_t:
        row=bloco_ok(row, "Tanques (Reg. 1310): nenhum valor negativo detectado")
    else:
        row=bloco_err(row, f"Tanques: {len(todos_neg_t)} valor(es) negativo(s) encontrado(s)")
        for x in todos_neg_t:
            row=detalhe(row, f"Tanque {x['tanque']}  |  Data: {x['data'].strftime('%d/%m/%Y') if x.get('data') else 'N/D'}  |  Campo: {x['campo']}  |  Valor: {x['valor']:,.3f}  |  Linha SPED: {x['linha']}")
            diag=DIAGNOSTICO_CAMPO.get(x['campo'],"Valor negativo inesperado neste campo")
            row=detalhe(row, f"     ℹ️  {diag}", cor_bg="FCE4D6", cor_fg="7D4200")

    if not todos_neg_b:
        row=bloco_ok(row, "Bicos (Reg. 1320): nenhum valor negativo detectado")
    else:
        row=bloco_err(row, f"Bicos: {len(todos_neg_b)} valor(es) negativo(s) encontrado(s)")
        for x in todos_neg_b:
            row=detalhe(row, f"Bico {x['bico']}  |  Data: {x['data'].strftime('%d/%m/%Y') if x.get('data') else 'N/D'}  |  Campo: {x['campo']}  |  Valor: {x['valor']:,.3f}  |  Linha SPED: {x['linha']}")
            diag=DIAGNOSTICO_CAMPO.get(x['campo'],"Valor negativo inesperado neste campo")
            row=detalhe(row, f"     ℹ️  {diag}", cor_bg="FCE4D6", cor_fg="7D4200")
    row+=1

    # 4. VERSÃO E CAPACIDADE
    _subtit(ws,row,f"4.  VERSÃO DO SPED E CAPACIDADE DOS TANQUES  –  {vc_mai['periodo']}",N); row+=1
    versao_txt = (f"✅  Versão do SPED: {vc_mai['versao']} — correta (obrigatório: {VERSAO_OBRIGATORIA})"
                  if vc_mai["versao_ok"]
                  else f"❌  Versão do SPED: {vc_mai['versao']} — incorreta! Obrigatório: {VERSAO_OBRIGATORIA}. Pode causar rejeição no validador.")
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=N)
    c=ws.cell(row=row,column=1,value=versao_txt)
    bg=C_VERDE_BG if vc_mai["versao_ok"] else C_VERM_BG
    fg=C_VERDE_FG if vc_mai["versao_ok"] else C_VERM_FG
    c.font=Font(name="Arial",bold=True,size=10,color=fg)
    c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(horizontal="left",vertical="center"); c.border=_brd()
    ws.row_dimensions[row].height=18; row+=1

    n_cap_ok  = sum(1 for t in vc_mai["tanques"] if t["status"]=="✅ OK")
    n_cap_err = sum(1 for t in vc_mai["tanques"] if t["status"]!="✅ OK")
    if n_cap_err==0:
        row=bloco_ok(row, f"Capacidade dos tanques: todos os {n_cap_ok} tanques declarados corretamente  |  "
                     + "  ".join(f"Tanque {t['tanque']}: {t['obs']}" for t in vc_mai["tanques"]))
    else:
        row=bloco_err(row, f"Capacidade: {n_cap_err} tanque(s) com problema")
        for t in vc_mai["tanques"]:
            bg2=_row_bg(t["status"]) or C_AMAR_BG
            fg2=C_VERM_FG if "❌" in t["status"] else C_AMAR_FG
            row=detalhe(row, f"Tanque {t['tanque']}  |  {t['status']}  |  {t['obs']}", cor_bg=bg2, cor_fg=fg2)

    # 5. CADASTRO LMC (bombas, lacres, bico-tanque)
    _subtit(ws,row,"5.  CADASTRO LMC  –  BOMBAS, LACRES E VÍNCULOS BICO-TANQUE",N); row+=1

    if cad_atu:
        # Bombas
        if cad_atu["bombas_ok"]:
            row=bloco_ok(row, f"Bombas (Reg. 1350): {cad_atu['n_bombas']} bomba(s) cadastrada(s)")
        else:
            row=bloco_err(row, "Bombas (Reg. 1350): nenhuma bomba declarada — incluir antes de enviar o SPED")

        # Lacres
        if cad_atu["afericoes_ok"]:
            row=bloco_ok(row, f"Lacres (Reg. 1360): {cad_atu['n_afericoes']} lacre(s) declarado(s)")
        else:
            row=bloco_err(row, "Lacres (Reg. 1360): nenhum lacre declarado — incluir número e data de instalação de cada lacre")

        # Bico-Tanque
        if cad_atu["lacres_ok"]:
            row=bloco_ok(row, f"Vínculos Bico-Tanque (Reg. 1370): {cad_atu['n_lacres']} vínculo(s) declarado(s)")
        else:
            row=bloco_err(row, "Vínculos Bico-Tanque (Reg. 1370): nenhum vínculo declarado — informar qual bico está ligado a qual tanque")
    else:
        row=bloco_ok(row, "Cadastro LMC: não verificado (dados não disponíveis)")

    ws.column_dimensions["A"].width=90
    for i in range(2,N+1):
        ws.column_dimensions[get_column_letter(i)].width=0.1

# ── ABA CONFRONTO MENSAL ───────────────────────────────────────────────────────
def aba_mensal(ws, conf_m, info_ant, info_atu):
    ws.sheet_view.showGridLines = False
    r = 1

    def fmt_comp(dt):
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try:
            mes = int(dt[2:4]); ano = dt[4:]
            return f"{meses[mes-1]}/{ano}"
        except:
            return dt

    comp_ant = fmt_comp(info_ant["info"].get("dt_fin", "")) or "—"
    comp_atu = fmt_comp(info_atu["info"].get("dt_fin", ""))

    if not conf_m["tanques"] and not conf_m["bicos"]:
        _titulo(ws, r, "CONFRONTO ENTRE MESES", 6, sz=13); r += 2
        cel = ws.cell(row=r, column=1,
            value="⚠️  SPED da competência anterior não foi enviado. "
                  "Não é possível confrontar o fechamento do mês anterior com a abertura "
                  "do mês atual. Consulte as demais abas para a conferência da competência atual.")
        cel.font = Font(name="Arial", size=11, bold=True, color=C_AMAR_FG)
        cel.fill = PatternFill("solid", start_color=C_AMAR_BG)
        cel.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=6)
        ws.row_dimensions[r].height=60
        ws.column_dimensions["A"].width=90
        return

    for secao, lista, tipo, lbl_fech_ant, lbl_aber_atu, lbl_fech_atu in [
        ("TANQUES – Confronto Competência Anterior × Atual (Reg. 1310)",
         conf_m["tanques"], "Tanque",
         f"Est. Fechamento\n{comp_ant} (L)",
         f"Est. Abertura\n{comp_atu} (L)",
         f"Est. Fechamento\n{comp_atu} (L)"),
        ("BICOS – Confronto Competência Anterior × Atual (Reg. 1320)",
         conf_m["bicos"], "Bico",
         f"Enc. Fechamento\n{comp_ant}",
         f"Enc. Abertura\n{comp_atu}",
         f"Enc. Fechamento\n{comp_atu}"),
    ]:
        _titulo(ws, r, secao, 6); r += 1
        for i, h in enumerate([tipo, lbl_fech_ant, lbl_aber_atu, "Diferença", "Status", lbl_fech_atu], 1):
            _ch(ws, r, i, h)
        ws.row_dimensions[r].height = 34; r += 1

        for x in lista:
            bg = _row_bg(x["status"])
            _dc(ws, r, 1, f"{tipo} {x['id']}", bg=bg)
            _dc(ws, r, 2, x["fech"],     NF, bg=bg)
            _dc(ws, r, 3, x["aber"],     NF, bg=bg)
            _dc(ws, r, 4, x["dif"],      NF, bg=bg)
            _sc(ws, r, 5, x["status"])
            _dc(ws, r, 6, x.get("fech_atu"), NF, bg=bg)
            ws.row_dimensions[r].height = 15; r += 1
        r += 2

    for i, w in enumerate([14, 22, 22, 16, 22, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def aba_diario(ws, d_mai):
    ws.sheet_view.showGridLines=False; r=1
    for secao,lista,id_f,cols in [
        ("TANQUES – Comparativo Diário Competência Atual (Reg. 1310)",
         d_mai["tanques"],"tanque",
         ["Tanque","Data Fech.","Est. Fechamento (L)","Data Aber.","Est. Abertura (L)","Diferença (L)","Status"]),
        ("BICOS – Comparativo Diário Competência Atual (Reg. 1320)",
         d_mai["bicos"],"bico",
         ["Bico","Data Fech.","Enc. Fechamento","Data Aber.","Enc. Abertura","Diferença","Status"]),
    ]:
        _titulo(ws,r,secao,8); r+=1
        for i,h in enumerate(cols,1): _ch(ws,r,i,h)
        ws.row_dimensions[r].height=30; r+=1
        tipo=cols[0]  # "Tanque" ou "Bico"
        for x in lista:
            bg=_row_bg(x["status"])
            _dc(ws,r,1,f"{tipo} {x[id_f]}",bg=bg)
            _dc(ws,r,2,x["dia_fech"].strftime("%d/%m/%Y"),bg=bg)
            _dc(ws,r,3,x["fech"],NF,bg=bg)
            _dc(ws,r,4,x["dia_aber"].strftime("%d/%m/%Y"),bg=bg)
            _dc(ws,r,5,x["aber"],NF,bg=bg)
            _dc(ws,r,6,x["dif"],NF,bg=bg)
            _sc(ws,r,7,x["status"])
            ws.row_dimensions[r].height=15; r+=1
        r+=2
    for i,w in enumerate([14,14,22,14,22,16,22],1):
        ws.column_dimensions[get_column_letter(i)].width=w

# ══════════════════════════════════════════════════════════════════════════════
# MAIN

def ler_sped_bytes(data):
    """Versão para receber bytes (upload web) ao invés de caminho de arquivo."""
    text = data.decode("latin-1", errors="replace")
    tanques={}; bicos={}; info={}
    data_atual=None; vals_1300={}

    for linha in text.splitlines():
        c=linha.strip().split("|")
        if len(c)<2: continue
        tp=c[1]

        if tp=="0000":
            info={"versao":c[2].strip() if len(c)>2 else "",
                  "razao": c[6].strip() if len(c)>6 else "",
                  "cnpj":  c[7].strip() if len(c)>7 else "",
                  "dt_ini":c[4].strip() if len(c)>4 else "",
                  "dt_fin":c[5].strip() if len(c)>5 else ""}

        elif tp=="1300":
            data_atual=_dt(c[3]) if len(c)>3 else None
            if not data_atual: continue
            vals_1300={
                "data":data_atual,
                "est_abert":_fl(c[4]),"entrada":_fl(c[5]),"saida":_fl(c[7]),
                "evap":_fl(c[9]) if len(c)>9 else 0.0,
                "ajuste":_fl(c[10]) if len(c)>10 else 0.0,
                "est_fech":_fl(c[11]) if len(c)>11 else 0.0,
            }

        elif tp=="1310":
            if not vals_1300: continue
            t=_nid(c[2])
            tem_campos_proprios = len(c) > 10 and c[3].strip() != ""
            if tem_campos_proprios:
                try:
                    est_abert_1310 = _fl(c[3])
                    entrada_1310   = _fl(c[4]) if len(c)>4 else 0.0
                    saida_1310     = _fl(c[6]) if len(c)>6 else 0.0
                    evap_1310      = _fl(c[8]) if len(c)>8 else 0.0
                    ajuste_1310    = _fl(c[9]) if len(c)>9 else 0.0
                    est_fech_1310  = _fl(c[10]) if len(c)>10 else 0.0
                    cap            = _fl(c[11]) if len(c)>11 and c[11].strip() else None
                except Exception:
                    tem_campos_proprios = False

            if tem_campos_proprios:
                key=(t,vals_1300["data"])
                tanques[key]={
                    "tanque":t,"data":vals_1300["data"],
                    "est_abert":est_abert_1310,"entrada":entrada_1310,
                    "saida":saida_1310,"evap":evap_1310,
                    "ajuste":ajuste_1310,"est_fech":est_fech_1310,
                    "capacidade":cap,
                }
            else:
                cap=_fl(c[11]) if len(c)>11 and c[11].strip() else None
                key=(t,vals_1300["data"])
                tanques[key]={
                    "tanque":t,"data":vals_1300["data"],
                    "est_abert":vals_1300["est_abert"],"entrada":vals_1300["entrada"],
                    "saida":vals_1300["saida"],"evap":vals_1300["evap"],
                    "ajuste":vals_1300["ajuste"],"est_fech":vals_1300["est_fech"],
                    "capacidade":cap,
                }

        elif tp=="1320":
            b=_nid(c[2])
            if not data_atual: continue
            bicos[(b,data_atual)]={
                "bico":b,"data":data_atual,
                "enc_abert":_fl(c[9]) if len(c)>9 else 0.0,
                "enc_fech": _fl(c[8]) if len(c)>8 else 0.0,
            }

    return {"info":info,"tanques":tanques,"bicos":bicos}



# ── CAMPOS DO GUIA LMC ───────────────────────────────────────────────────────
VERSAO_OBRIGATORIA = "020"

PRODUTOS_COMBUSTIVEL = {
    "000001": "GASOLINA COMUM",
    "000002": "DIESEL S10",
    "000003": "DIESEL S500",
    "000004": "ETANOL COMUM",
    "000005": "GASOLINA ADITIVADA",
    "000006": "DIESEL ADITIVADO",
}

def verificar_cadastro_lmc(data):
    """
    Verifica se tanques (1310), bicos (1320), bombas (1350),
    aferições (1360) e lacres (1370) estão preenchidos no SPED.
    Retorna um dict com o resultado de cada verificação.
    """
    text = data.decode("latin-1", errors="replace")
    linhas = text.splitlines()

    tanques  = {}   # {num: {cap, dias_declarados}}
    bicos    = {}   # {num: {dias_declarados}}
    bombas   = []   # lista de {serie, fabricante, modelo}
    afericoes= []   # lista de {num, data}
    lacres   = []   # lista de {lacre, seq}

    for l in linhas:
        c = l.strip().split("|")
        if len(c) < 2: continue
        tp = c[1]

        if tp == "1310":
            num = _nid(c[2]) if len(c)>2 else ""
            cap = c[11].strip() if len(c)>11 else ""
            if num not in tanques:
                tanques[num] = {"capacidade": cap, "dias": 0}
            tanques[num]["dias"] += 1
            if cap and not tanques[num]["capacidade"]:
                tanques[num]["capacidade"] = cap

        elif tp == "1320":
            num = _nid(c[2]) if len(c)>2 else ""
            if num not in bicos:
                bicos[num] = {"dias": 0}
            bicos[num]["dias"] += 1

        elif tp == "1350":
            bombas.append({
                "serie":      c[2].strip() if len(c)>2 else "",
                "fabricante": c[3].strip() if len(c)>3 else "",
                "modelo":     c[4].strip() if len(c)>4 else "",
            })

        elif tp == "1360":
            afericoes.append({
                "numero": c[2].strip() if len(c)>2 else "",
                "data":   c[3].strip() if len(c)>3 else "",
            })

        elif tp == "1370":
            # 1370: [2]=Nº Bico, [3]=Cód.Produto, [4]=Nº Tanque
            lacres.append({
                "bico":    c[2].strip() if len(c)>2 else "",
                "produto": c[3].strip() if len(c)>3 else "",
                "tanque":  c[4].strip() if len(c)>4 else "",
            })

    # ── Resultado ─────────────────────────────────────────────────────────────
    # Tanques: OK se declarados, ❌ se nenhum
    tanques_ok = len(tanques) > 0
    # Capacidade: verificar se todos têm capacidade declarada
    tanques_sem_cap = [t for t, v in tanques.items() if not v["capacidade"] or v["capacidade"] in ("", "0")]

    # Bicos: OK se declarados
    bicos_ok = len(bicos) > 0

    # Bombas (1350): OK se declaradas
    bombas_ok = len(bombas) > 0

    # Aferições (1360): OK se declaradas
    afericoes_ok = len(afericoes) > 0

    # Lacres (1370): OK se declarados
    lacres_ok = len(lacres) > 0

    return {
        "tanques":         sorted(tanques.keys(), key=_sk),
        "tanques_ok":      tanques_ok,
        "tanques_sem_cap": sorted(tanques_sem_cap, key=_sk),
        "bicos":           sorted(bicos.keys(), key=_sk),
        "bicos_ok":        bicos_ok,
        "bombas":          bombas,
        "bombas_ok":       bombas_ok,
        "afericoes":       afericoes,
        "afericoes_ok":    afericoes_ok,
        "bicos_lacres":    lacres,   # 1370: bico → produto → tanque
        "lacres_ok":       lacres_ok,
        "n_tanques":       len(tanques),
        "n_bicos":         len(bicos),
        "n_bombas":        len(bombas),
        "n_afericoes":     len(afericoes),
        "n_lacres":        len(lacres),
    }


def verificar_negativos_bytes(data):
    """
    Verifica valores negativos nos registros 1300/1310/1320.
    O número real do tanque vem do campo[2] do registro 1310 filho,
    NÃO do campo[2] do 1300 pai (que é o código do produto no SPED).
    """
    text = data.decode("latin-1", errors="replace")
    neg_t=[]; neg_b=[]; da=None
    vals_1300={}   # guarda campos do 1300 atual para usar no 1310
    tanque_atual=None  # número do tanque resolvido via 1310

    for n, linha in enumerate(text.splitlines(), 1):
        c = linha.strip().split("|")
        if len(c)<2: continue
        tp=c[1]

        if tp=="1300":
            da=_dt(c[3]) if len(c)>3 else None
            # Guardar campos do 1300 para verificação posterior via 1310
            vals_1300 = {"data": da, "linha": n, "campos": c}
            tanque_atual = None  # será resolvido no 1310

        elif tp=="1310":
            # O 1310 traz o número real do tanque em c[2]
            tanque_atual = _nid(c[2])

            # Verificar campos do 1310 (que têm os valores reais do tanque)
            tem_campos_proprios = len(c) > 10 and c[3].strip() != ""
            if tem_campos_proprios:
                # Usar campos do próprio 1310
                campos_check = {
                    3:  "Est. Abertura",
                    6:  "Saída",
                    8:  "Evaporação",
                    9:  "Ajuste",
                    10: "Est. Fechamento Final",
                }
                for idx, nome in campos_check.items():
                    if idx >= len(c): continue
                    try:
                        v = _fl(c[idx])
                        if v is not None and v < 0:
                            neg_t.append({"tanque": tanque_atual, "data": da,
                                          "campo": nome, "valor": v, "linha": n})
                    except: pass
            else:
                # Herdar do 1300 pai — verificar campos do 1300 usando número do 1310
                c1300 = vals_1300.get("campos", [])
                for idx, nome in CAMPOS_1300.items():
                    if idx >= len(c1300): continue
                    try:
                        v = _fl(c1300[idx])
                        if v is not None and v < 0:
                            neg_t.append({"tanque": tanque_atual, "data": da,
                                          "campo": nome, "valor": v,
                                          "linha": vals_1300.get("linha", n)})
                    except: pass

        elif tp=="1320":
            bico=_nid(c[2])
            for idx,nome in CAMPOS_1320.items():
                if idx>=len(c): continue
                try:
                    v=_fl(c[idx])
                    if v is not None and v<0:
                        neg_b.append({"bico":bico,"data":da,"campo":nome,"valor":v,"linha":n})
                except: pass

    return {"tanques": neg_t, "bicos": neg_b}

# Diagnóstico por campo negativo
DIAGNOSTICO_CAMPO = {
    "Est. Abertura":          "Estoque de abertura negativo — erro de lançamento ou fechamento anterior incorreto",
    "Entrada":                "Entrada de combustível negativa — possível estorno de nota fiscal ou lançamento incorreto",
    "Est. Aber. Pós Entrada": "Estoque após entrada negativo — inconsistência entre abertura e entrada lançadas",
    "Saída":                  "Saída negativa — possível estorno ou correção de venda lançada indevidamente",
    "Est. Fech. Pré Ajuste":  "Estoque pré-ajuste negativo — vendas ou perdas superiores ao estoque disponível",
    "Evaporação":             "Evaporação negativa — valor inválido; evaporação deve ser sempre positiva ou zero",
    "Ajuste":                 "Ajuste negativo — perda de inventário ou correção de estoque para menor",
    "Est. Fechamento Final":  "Estoque de fechamento negativo — saldo final abaixo de zero, impossível fisicamente",
    "Enc. Fechamento":        "Encerrante de fechamento negativo — erro de leitura do bico ou lançamento incorreto",
    "Enc. Abertura":          "Encerrante de abertura negativo — valor inválido; encerrante é sempre crescente",
    "Volume Vendido":         "Volume vendido negativo — possível estorno ou erro de registro no bico",
    "Diferença Encerrante":   "Diferença de encerrante negativa — bico apresentou recuo, possível adulteração ou falha",
}


def aba_dac(ws, conf_dac, info_atu):
    """Aba de confronto DAC × SPED da competência atual."""
    ws.sheet_view.showGridLines = False
    r = 1

    def fmt_comp(dt):
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try: return f"{meses[int(dt[2:4])-1]}/{dt[4:]}"
        except: return dt

    comp_atu = fmt_comp(info_atu["info"].get("dt_fin",""))
    comp_dac = conf_dac.get("competencia", comp_atu)

    _titulo(ws, r, f"CONFRONTO DAC × SPED  –  Competência {comp_atu}", 8, sz=13)
    ws.row_dimensions[r].height = 30; r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(row=r, column=1,
        value=f"DAC competência: {comp_dac}  |  SPED competência: {comp_atu}  |  "
              f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(name="Arial", size=9, italic=True, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 16; r += 2

    # ── TANQUES ──────────────────────────────────────────────────────────────
    _subtit(ws, r, f"TANQUES (Reg. 1310)  –  Estoque Inicial e Final {comp_atu}", 8); r += 1
    for i, h in enumerate([
        "Tanque", "Produto",
        "Est. Inicial\nDAC (L)", "Est. Inicial\nSPED (L)", "Dif. Inicial",
        "Est. Final\nDAC (L)",   "Est. Final\nSPED (L)",   "Dif. Final",
    ], 1): _ch(ws, r, i, h, bg=C_CINZA, fg=C_AZUL_ESC)
    ws.row_dimensions[r].height = 30; r += 1

    for x in conf_dac.get("tanques", []):
        st_i = x["status_ini"]; st_f = x["status_fin"]
        bg = (C_VERM_BG if "DIVERGÊNCIA" in (st_i+st_f) else
              C_AMAR_BG if "AUSENTE"     in (st_i+st_f) else C_VERDE_BG)
        _dc(ws,r,1, f"Tanque {x['id']}", bg=bg)
        _dc(ws,r,2, x.get("produto",""), bg=bg)
        _dc(ws,r,3, x["ei_dac"],  NF, bg=bg)
        _dc(ws,r,4, x["ei_sped"], NF, bg=bg)
        _dc(ws,r,5, x["dif_ini"], NF, bg=bg)
        _dc(ws,r,6, x["ef_dac"],  NF, bg=bg)
        _dc(ws,r,7, x["ef_sped"], NF, bg=bg)
        _dc(ws,r,8, x["dif_fin"], NF, bg=bg)
        ws.row_dimensions[r].height = 15; r += 1
    r += 1

    # ── BICOS ─────────────────────────────────────────────────────────────────
    _subtit(ws, r, f"BICOS (Reg. 1320)  –  Encerrante Inicial e Final {comp_atu}", 7); r += 1
    for i, h in enumerate([
        "Bico",
        "Enc. Inicial\nDAC", "Enc. Inicial\nSPED", "Dif. Inicial",
        "Enc. Final\nDAC",   "Enc. Final\nSPED",   "Dif. Final",
    ], 1): _ch(ws, r, i, h, bg=C_CINZA, fg=C_AZUL_ESC)
    ws.row_dimensions[r].height = 30; r += 1

    for x in conf_dac.get("bicos", []):
        st_i = x["status_ini"]; st_f = x["status_fin"]
        bg = (C_VERM_BG if "DIVERGÊNCIA" in (st_i+st_f) else
              C_AMAR_BG if "AUSENTE"     in (st_i+st_f) else C_VERDE_BG)
        _dc(ws,r,1, f"Bico {x['id']}", bg=bg)
        _dc(ws,r,2, x["ei_dac"],  NF, bg=bg)
        _dc(ws,r,3, x["ei_sped"], NF, bg=bg)
        _dc(ws,r,4, x["dif_ini"], NF, bg=bg)
        _dc(ws,r,5, x["ef_dac"],  NF, bg=bg)
        _dc(ws,r,6, x["ef_sped"], NF, bg=bg)
        _dc(ws,r,7, x["dif_fin"], NF, bg=bg)
        ws.row_dimensions[r].height = 15; r += 1

    for i, w in enumerate([12, 22, 18, 18, 16, 18, 18, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def aba_dac_sped(ws, d_atu, info_atu):
    """Gera um DAC de acompanhamento a partir dos dados do SPED da competência atual."""
    ws.sheet_view.showGridLines = False
    r = 1
    N = 9

    def fmt_comp(dt):
        meses=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try: return f"{meses[int(dt[2:4])-1]}/{dt[4:]}"
        except: return dt

    info  = info_atu["info"]
    comp  = fmt_comp(info.get("dt_fin",""))
    razao = info.get("razao","")
    cnpj  = info.get("cnpj","")
    dt_ini= info.get("dt_ini","")
    dt_fin= info.get("dt_fin","")

    _titulo(ws, r, "DAC – DOCUMENTO DE ACOMPANHAMENTO DE COMBUSTÍVEIS (gerado pelo SPED)", N, sz=12)
    ws.row_dimensions[r].height = 28; r += 1

    for texto_a, texto_e in [
        (f"Empresa: {razao}", f"CNPJ: {cnpj}"),
        (f"Período: {dt_ini[:2]}/{dt_ini[2:4]}/{dt_ini[4:]} a {dt_fin[:2]}/{dt_fin[2:4]}/{dt_fin[4:]}", "Fonte: SPED Fiscal – Reg. 1300 / 1320"),
        (f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", "⚠️ Valores baseados no escritural do SPED (sem medição física)"),
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c1 = ws.cell(row=r, column=1, value=texto_a)
        c1.font = Font(name="Arial", size=10, bold=True)
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=N)
        c2 = ws.cell(row=r, column=6, value=texto_e)
        c2.font = Font(name="Arial", size=10, italic=True, color="595959")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[r].height = 16; r += 1
    r += 1

    # ── POSIÇÃO DOS TANQUES ───────────────────────────────────────────────────
    _subtit(ws, r, f"POSIÇÃO DOS TANQUES  –  {comp}", N); r += 1
    hdrs_t = ["Tanque","Produto","Est. Inicial (L)","Recebimento (L)","Venda (L)",
              "Evaporação (L)","Perda / Ganho (L)","Est. Final (L)","Variação (L)"]
    for i,h in enumerate(hdrs_t,1): _ch(ws,r,i,h)
    ws.row_dimensions[r].height = 28; r += 1

    for t in sorted(set(tt for (tt,_) in d_atu["tanques"]), key=_sk):
        dias = sorted(dt for (tt,dt) in d_atu["tanques"] if tt==t)
        est_ini  = d_atu["tanques"][(t,dias[0])]["est_abert"]
        est_fin  = d_atu["tanques"][(t,dias[-1])]["est_fech"]
        total_rec= round(sum(d_atu["tanques"][(t,dt)]["entrada"] for dt in dias),3)
        total_sai= round(sum(d_atu["tanques"][(t,dt)]["saida"]   for dt in dias),3)
        total_evap=round(sum(d_atu["tanques"][(t,dt)]["evap"]    for dt in dias),3)
        total_aj = round(sum(d_atu["tanques"][(t,dt)]["ajuste"]  for dt in dias),3)
        variacao = round(est_fin - est_ini, 3)

        cap = d_atu["tanques"].get((t,dias[0]),{}).get("capacidade")
        produto = f"{cap:,.0f} L" if cap else "—"

        if total_aj < 0:
            bg_aj = C_VERM_BG; fg_aj = C_VERM_FG
        elif total_aj > 0:
            bg_aj = C_VERDE_BG; fg_aj = C_VERDE_FG
        else:
            bg_aj = None; fg_aj = "000000"

        _dc(ws,r,1,f"Tanque {t}")
        _dc(ws,r,2,produto)
        _dc(ws,r,3,est_ini,   NF)
        _dc(ws,r,4,total_rec, NF, bg="D6E4F0" if total_rec>0 else None)
        _dc(ws,r,5,total_sai, NF)
        _dc(ws,r,6,total_evap,NF)
        cel_aj = ws.cell(row=r,column=7,value=total_aj)
        cel_aj.number_format = NF
        cel_aj.font = Font(name="Arial",size=10,bold=(total_aj!=0),color=fg_aj)
        cel_aj.alignment = Alignment(horizontal="center",vertical="center")
        cel_aj.border = _brd()
        if bg_aj: cel_aj.fill = PatternFill("solid",start_color=bg_aj)
        _dc(ws,r,8,est_fin,   NF)
        cel_v = ws.cell(row=r,column=9,value=variacao)
        cel_v.number_format = NF
        cel_v.font = Font(name="Arial",size=10,
                          color=C_VERDE_FG if variacao>=0 else C_VERM_FG,bold=True)
        cel_v.alignment = Alignment(horizontal="center",vertical="center")
        cel_v.border = _brd()
        cel_v.fill = PatternFill("solid", start_color=C_VERDE_BG if variacao>=0 else C_VERM_BG)
        ws.row_dimensions[r].height = 16; r += 1
    r += 2

    # ── POSIÇÃO DOS BICOS ─────────────────────────────────────────────────────
    _subtit(ws, r, f"POSIÇÃO DOS BICOS  –  {comp}", N); r += 1
    hdrs_b = ["Bico","Enc. Inicial","Enc. Final","Litros Vendidos"]
    for i,h in enumerate(hdrs_b,1): _ch(ws,r,i,h,bg=C_AZUL_ESC)
    ws.row_dimensions[r].height = 24; r += 1

    total_litros = 0
    for b in sorted(set(bb for (bb,_) in d_atu["bicos"]), key=_sk):
        dias = sorted(dt for (bb,dt) in d_atu["bicos"] if bb==b)
        enc_ini = d_atu["bicos"][(b,dias[0])]["enc_abert"]
        enc_fin = d_atu["bicos"][(b,dias[-1])]["enc_fech"]
        litros  = round(enc_fin - enc_ini, 3)
        total_litros += litros

        bg = C_VERDE_BG if litros>0 else C_AMAR_BG
        _dc(ws,r,1,f"Bico {b}")
        _dc(ws,r,2,enc_ini,NF)
        _dc(ws,r,3,enc_fin,NF)
        _dc(ws,r,4,litros, NF, bg=bg)
        ws.row_dimensions[r].height = 15; r += 1

    for i,v in enumerate(["TOTAL","","",round(total_litros,3)],1):
        c=ws.cell(row=r,column=i,value=v)
        c.font=Font(name="Arial",bold=True,size=10)
        c.fill=PatternFill("solid",start_color=C_CINZA)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=_brd()
        if i==4: c.number_format=NF
    ws.row_dimensions[r].height=18; r+=1

    for i,w in enumerate([12,14,18,18,18,16,14,18,16],1):
        ws.column_dimensions[get_column_letter(i)].width=w


def _fmt(v, decimais=3):
    """Formata número com segurança, retornando '—' se None."""
    if v is None: return "—"
    try: return f"{v:,.{decimais}f}"
    except: return str(v)

def aba_relatorio_cliente(ws, conf_m, d_mai, neg_abr, neg_mai, vc_mai, conf_dac, info_ant, info_atu, cad_atu=None):
    """Aba de notificação de divergências para envio ao cliente — layout limpo."""
    ws.sheet_view.showGridLines = False

    N = 6
    COLS = {"A":22,"B":18,"C":18,"D":18,"E":18,"F":18}
    for letra, w in COLS.items():
        ws.column_dimensions[letra].width = w

    ia = info_ant.get("info", {}); iu = info_atu.get("info", {})
    razao    = ia.get("razao","") or iu.get("razao","")
    cnpj_raw = ia.get("cnpj","")  or iu.get("cnpj","")
    cnpj_fmt = (f"{cnpj_raw[:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}"
                f"/{cnpj_raw[8:12]}-{cnpj_raw[12:]}") if len(cnpj_raw)==14 else cnpj_raw

    def fmt_comp(dt):
        meses=["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
        try: return f"{meses[int(dt[2:4])-1]}/{dt[4:]}"
        except: return dt or "—"

    comp_atu = fmt_comp(iu.get("dt_fin",""))
    comp_ant = fmt_comp(ia.get("dt_fin","")) if ia.get("dt_fin") else None

    def aplic_brd(r, c1, c2):
        for col in range(c1, c2+1):
            ws.cell(row=r, column=col).border = _brd()

    def tit(r, texto):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        c = ws.cell(row=r, column=1, value=texto)
        c.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=C_AZUL_ESC)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 28
        aplic_brd(r, 1, N)

    def subt(r, texto):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        c = ws.cell(row=r, column=1, value=texto)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=C_AZUL_MED)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        aplic_brd(r, 1, N)

    def cab(r, headers, largs):
        col = 1
        for h, larg in zip(headers, largs):
            if larg > 1:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+larg-1)
            c = ws.cell(row=r, column=col, value=h)
            c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="831040")
            c.alignment = Alignment(horizontal="center", vertical="center")
            aplic_brd(r, col, col+larg-1)
            col += larg
        ws.row_dimensions[r].height = 20

    def lin(r, valores, largs):
        col = 1
        for i, (v, larg) in enumerate(zip(valores, largs)):
            if larg > 1:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+larg-1)
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Arial", size=9, color="1a2340")
            c.alignment = Alignment(horizontal="left" if i==0 else "center", vertical="center")
            aplic_brd(r, col, col+larg-1)
            col += larg
        ws.row_dimensions[r].height = 16

    def ok(r, texto):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        c = ws.cell(row=r, column=1, value=f"✅  {texto}")
        c.font = Font(name="Arial", size=9, color=C_VERDE_FG)
        c.fill = PatternFill("solid", start_color=C_VERDE_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 18
        aplic_brd(r, 1, N)

    def fmt_dt(dt):
        try: return dt.strftime("%d/%m/%Y")
        except: return str(dt)

    row = 1
    tit(row, "NOTIFICAÇÃO DE DIVERGÊNCIAS – LIVRO DE MOVIMENTAÇÃO DE COMBUSTÍVEIS")
    _inserir_logo(ws, f"{get_column_letter(N)}{row}", altura=34)
    row += 1

    for label, valor in [
        ("Empresa:", razao),
        ("CNPJ:", cnpj_fmt),
        ("Competência:", comp_atu),
        ("Data:", datetime.now().strftime("%d/%m/%Y")),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name="Arial", size=9, bold=True, color="555555")
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=N)
        c2 = ws.cell(row=row, column=3, value=valor)
        c2.font = Font(name="Arial", size=9, bold=(label=="Empresa:"), color="1a2340")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    intro = ws.cell(row=row, column=1,
        value=("A seguir, apresentamos as divergências identificadas na conferência do LMC. "
               "Solicitamos a verificação e o encaminhamento dos arquivos corrigidos. "
               "Inconsistências no LMC podem acarretar penalidades perante a SEFAZ e a ANP."))
    intro.font = Font(name="Arial", size=9, italic=True, color="666666")
    intro.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 2

    # 1. Confronto entre meses
    divs_m = [x for x in conf_m["tanques"] if x["status"]!="✅ OK"]
    divs_m += [x for x in conf_m["bicos"]  if x["status"]!="✅ OK"]
    subt(row, f"1.  CONFRONTO ENTRE MESES  ({comp_ant} × {comp_atu})" if comp_ant
              else f"1.  CONFRONTO ENTRE MESES  – {comp_atu}")
    row += 1
    if not divs_m:
        ok(row, "Fechamento do mês anterior e abertura atual conferidos — sem divergências."); row += 1
    else:
        lm = [2,1,2,1]
        cab(row, ["Tanque/Bico","Fech. Anterior (L)","Aber. Atual (L)","Diferença (L)"], lm); row += 1
        for x in divs_m:
            lin(row, [f"Tanque {x['id']}", _fmt(x.get("fech")), _fmt(x.get("aber")), _fmt(x.get("dif"))], lm); row += 1
    row += 1

    # 2. Consistência diária
    divs_d = [x for x in d_mai["tanques"] if x["status"]!="✅ OK"]
    divs_d += [x for x in d_mai["bicos"]  if x["status"]!="✅ OK"]
    subt(row, f"2.  CONSISTÊNCIA DIÁRIA – {comp_atu}"); row += 1
    if not divs_d:
        ok(row, "Todas as transições diárias (fechamento → abertura) conferidas — sem divergências."); row += 1
    else:
        ld = [1,1,1,1,1,1]
        cab(row, ["Tanque","Data Fech.","Fechamento (L)","Data Aber.","Abertura (L)","Diferença (L)"], ld); row += 1
        for x in divs_d:
            id_key = "tanque" if "tanque" in x else "bico"
            lin(row, [f"Tanque {x[id_key]}", fmt_dt(x["dia_fech"]), _fmt(x["fech"]),
                      fmt_dt(x["dia_aber"]), _fmt(x["aber"]), _fmt(x["dif"])], ld); row += 1
    row += 1

    # 3. ANP
    divs_anp = [t for t in conf_m["tanques"] if t.get("status_anp","") not in ("✅ DENTRO DO LIMITE","")]
    subt(row, "3.  LIMITE DE VARIAÇÃO ANP (0,6%)"); row += 1
    if not divs_anp:
        ok(row, "Todos os tanques dentro do limite de variação de 0,6% permitido pela ANP."); row += 1
    else:
        la = [2,1,1,1,1]
        cab(row, ["Tanque","Recebimento (L)","Variação (L)","Limite (L)","% Variação"], la); row += 1
        for t in divs_anp:
            lin(row, [f"Tanque {t['id']}", _fmt(t.get("total_rec") or 0, 0),
                      _fmt(t.get("diferenca_anp") or 0), _fmt(t.get("limite_anp") or 0, 0),
                      f"{t.get('pct_anp') or 0:.3f}%"], la); row += 1
    row += 1

    # 4. Estoque negativo
    todos_neg = neg_abr["tanques"]+neg_mai["tanques"]+neg_abr["bicos"]+neg_mai["bicos"]
    subt(row, "4.  ESTOQUE NEGATIVO"); row += 1
    if not todos_neg:
        ok(row, "Nenhum valor negativo detectado nos registros do SPED."); row += 1
    else:
        ln = [2,1,2,1]
        cab(row, ["Tanque/Bico","Data","Campo","Valor"], ln); row += 1
        for x in todos_neg:
            tipo = "Tanque" if "tanque" in x else "Bico"
            lin(row, [f"{tipo} {x.get('tanque') or x.get('bico')}",
                      fmt_dt(x["data"]) if x.get("data") else "—",
                      x.get("campo",""), _fmt(x.get("valor"))], ln); row += 1
    # ── 5. Cadastro LMC (bombas, lacres, bico-tanque) ───────────────────────
    if cad_atu:
        problemas_cad = []
        if not cad_atu["bombas_ok"]:
            problemas_cad.append("Bombas (Reg. 1350): nenhuma bomba declarada")
        if not cad_atu["afericoes_ok"]:
            problemas_cad.append("Lacres (Reg. 1360): nenhum lacre declarado")
        if not cad_atu["lacres_ok"]:
            problemas_cad.append("Vínculos Bico-Tanque (Reg. 1370): não declarados")

        if problemas_cad:
            subt(row, "5.  CADASTRO LMC  –  BOMBAS, LACRES E VÍNCULOS BICO-TANQUE"); row += 1
            ld_cad = [2, 4]
            cab(row, ["Item", "Pendência"], ld_cad); row += 1
            for prob in problemas_cad:
                item, desc = prob.split(": ", 1)
                lin(row, [item, desc], ld_cad); row += 1
            row += 1

    row += 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=N)
    rod = ws.cell(row=row, column=1,
        value=("Reforçamos a importância da regularização das divergências dentro do prazo. "
               "Solicitamos a gentileza de encaminhar os novos SPEDs fiscais com as devidas "
               "correções para que possamos concluir a conferência."))
    rod.font = Font(name="Arial", size=9, italic=True, color="666666")
    rod.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    aplic_brd(row, 1, N)



# ═════════════════════════════════════════════════════════════════════════════
# PAINEL ADMIN
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/admin")
@requer_perfil(PERFIL_ADMIN)
def admin_dashboard():
    stats = {
        "contabilidades": Contabilidade.query.filter_by(ativa=True).count(),
        "postos":         Posto.query.filter_by(ativo=True).count(),
        "usuarios":       Usuario.query.filter_by(ativo=True).count(),
        "relatorios":     Relatorio.query.count(),
    }
    postos_recentes = Posto.query.order_by(Posto.criado_em.desc()).limit(10).all()
    return render_template("admin/dashboard.html", stats=stats,
                           postos_recentes=postos_recentes, usuario=current_user)


@app.route("/admin/contabilidades")
@requer_perfil(PERFIL_ADMIN)
def admin_contabilidades():
    lista = Contabilidade.query.order_by(Contabilidade.nome).all()
    return render_template("admin/contabilidades.html", lista=lista, usuario=current_user)


@app.route("/admin/contabilidades/nova", methods=["GET","POST"])
@requer_perfil(PERFIL_ADMIN)
def admin_nova_contabilidade():
    if request.method == "POST":
        nome  = request.form.get("nome","").strip()
        cnpj  = re.sub(r"\D","", request.form.get("cnpj",""))
        email = request.form.get("email","").strip().lower()
        senha = request.form.get("senha","").strip()
        tel   = request.form.get("telefone","").strip()
        if not all([nome, cnpj, email, senha]):
            flash("Preencha todos os campos obrigatórios.", "danger")
        elif Contabilidade.query.filter_by(cnpj=cnpj).first():
            flash("CNPJ já cadastrado.", "danger")
        elif Usuario.query.filter_by(email=email).first():
            flash("E-mail já cadastrado.", "danger")
        else:
            cont = Contabilidade(nome=nome, cnpj=cnpj, email=email, telefone=tel)
            db.session.add(cont); db.session.flush()
            u = Usuario(email=email, nome=nome,
                        perfil=PERFIL_CONTABILIDADE, contabilidade_id=cont.id)
            u.set_senha(senha)
            db.session.add(u); db.session.commit()
            flash(f"Contabilidade '{nome}' criada com sucesso!", "success")
            return redirect(url_for("admin_contabilidades"))
    return render_template("admin/form_contabilidade.html", usuario=current_user)


@app.route("/admin/postos")
@requer_perfil(PERFIL_ADMIN)
def admin_postos():
    postos = Posto.query.order_by(Posto.razao_social).all()
    return render_template("admin/postos.html", postos=postos, usuario=current_user)


@app.route("/admin/postos/novo", methods=["GET","POST"])
@requer_perfil(PERFIL_ADMIN)
def admin_novo_posto():
    contabilidades = Contabilidade.query.filter_by(ativa=True).order_by(Contabilidade.nome).all()
    if request.method == "POST":
        cnpj     = re.sub(r"\D","", request.form.get("cnpj",""))
        razao    = request.form.get("razao_social","").strip()
        fantasia = request.form.get("nome_fantasia","").strip()
        email    = request.form.get("email","").strip().lower()
        senha    = request.form.get("senha","").strip()
        tel      = request.form.get("telefone","").strip()
        cidade   = request.form.get("cidade","").strip()
        estado   = request.form.get("estado","").strip().upper()
        plano    = request.form.get("plano", "mensal")
        cont_id  = request.form.get("contabilidade_id") or None
        if not all([cnpj, razao, email, senha]):
            flash("Preencha todos os campos obrigatórios.", "danger")
        elif Posto.query.filter_by(cnpj=cnpj).first():
            flash("CNPJ já cadastrado.", "danger")
        elif Usuario.query.filter_by(email=email).first():
            flash("E-mail já cadastrado.", "danger")
        else:
            posto = Posto(cnpj=cnpj, razao_social=razao, nome_fantasia=fantasia,
                          email=email, telefone=tel, cidade=cidade, estado=estado,
                          plano=plano, contabilidade_id=cont_id)
            db.session.add(posto); db.session.flush()
            u = Usuario(email=email, nome=razao,
                        perfil=PERFIL_POSTO, posto_id=posto.id)
            u.set_senha(senha)
            db.session.add(u); db.session.commit()
            flash(f"Posto '{razao}' criado com sucesso!", "success")
            return redirect(url_for("admin_postos"))
    return render_template("admin/form_posto.html",
                           contabilidades=contabilidades, usuario=current_user)


@app.route("/admin/posto/<int:posto_id>/toggle")
@requer_perfil(PERFIL_ADMIN)
def admin_toggle_posto(posto_id):
    posto = Posto.query.get_or_404(posto_id)
    posto.licenca_ativa = not posto.licenca_ativa
    db.session.commit()
    status = "ativado" if posto.licenca_ativa else "desativado"
    flash(f"Posto '{posto.razao_social}' {status}.", "success")
    return redirect(url_for("admin_postos"))


@app.route("/admin/usuarios")
@requer_perfil(PERFIL_ADMIN)
def admin_usuarios():
    usuarios = Usuario.query.order_by(Usuario.perfil, Usuario.nome).all()
    contabilidades = Contabilidade.query.filter_by(ativa=True).order_by(Contabilidade.nome).all()
    return render_template("admin/usuarios.html", usuarios=usuarios,
                           contabilidades=contabilidades, usuario=current_user)


@app.route("/admin/usuarios/novo", methods=["POST"])
@requer_perfil(PERFIL_ADMIN)
def admin_novo_usuario():
    nome     = request.form.get("nome", "").strip()
    email    = request.form.get("email", "").strip().lower()
    senha    = request.form.get("senha", "").strip()
    perfil   = request.form.get("perfil", "").strip()
    cont_id  = request.form.get("contabilidade_id") or None

    if not all([nome, email, senha, perfil]):
        flash("Preencha todos os campos obrigatórios.", "danger")
    elif perfil not in [PERFIL_ADMIN, PERFIL_CONTABILIDADE, PERFIL_POSTO]:
        flash("Perfil inválido.", "danger")
    elif Usuario.query.filter_by(email=email).first():
        flash("E-mail já cadastrado.", "danger")
    else:
        u = Usuario(email=email, nome=nome, perfil=perfil,
                    contabilidade_id=cont_id if perfil == PERFIL_CONTABILIDADE else None)
        u.set_senha(senha)
        db.session.add(u)
        db.session.commit()
        flash(f"Usuário '{nome}' criado com sucesso!", "success")

    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuario/<int:uid>/toggle")
@requer_perfil(PERFIL_ADMIN)
def admin_toggle_usuario(uid):
    u = Usuario.query.get_or_404(uid)
    if u.email == current_user.email:
        flash("Você não pode desativar sua própria conta.", "danger")
    else:
        u.ativo = not u.ativo
        db.session.commit()
        flash(f"Usuário '{u.nome}' {'ativado' if u.ativo else 'desativado'}.", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuario/<int:uid>/resetar_senha", methods=["POST"])
@requer_perfil(PERFIL_ADMIN)
def admin_resetar_senha(uid):
    u = Usuario.query.get_or_404(uid)
    nova = request.form.get("nova_senha", "").strip()
    if len(nova) < 6:
        flash("A senha deve ter ao menos 6 caracteres.", "danger")
    else:
        u.set_senha(nova)
        db.session.commit()
        flash(f"Senha de '{u.nome}' redefinida com sucesso!", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/relatorios")
@requer_perfil(PERFIL_ADMIN)
def admin_relatorios():
    rels = Relatorio.query.order_by(Relatorio.gerado_em.desc()).limit(200).all()
    return render_template("admin/relatorios.html", relatorios=rels, usuario=current_user)


# ═════════════════════════════════════════════════════════════════════════════
# PAINEL CONTABILIDADE
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/contabilidade/postos")
@requer_perfil(PERFIL_CONTABILIDADE, PERFIL_ADMIN)
def cont_postos():
    if current_user.is_admin:
        postos = Posto.query.order_by(Posto.razao_social).all()
    else:
        postos = Posto.query.filter_by(
            contabilidade_id=current_user.contabilidade_id
        ).order_by(Posto.razao_social).all()
    return render_template("contabilidade/postos.html", postos=postos, usuario=current_user)


@app.route("/contabilidade/posto/<int:posto_id>/relatorios")
@requer_perfil(PERFIL_CONTABILIDADE, PERFIL_ADMIN)
def cont_relatorios_posto(posto_id):
    posto = Posto.query.get_or_404(posto_id)
    # Contabilidade só vê postos vinculados a ela
    if current_user.is_contabilidade:
        if posto.contabilidade_id != current_user.contabilidade_id:
            abort(403)
    rels = Relatorio.query.filter_by(posto_id=posto_id)\
                          .order_by(Relatorio.gerado_em.desc()).all()
    return render_template("contabilidade/relatorios_posto.html",
                           posto=posto, relatorios=rels, usuario=current_user)


# ═════════════════════════════════════════════════════════════════════════════
# PAINEL POSTO (cliente final)
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/posto")
@requer_perfil(PERFIL_POSTO)
def posto_dashboard():
    posto = current_user.posto
    if not posto:
        flash("Posto não configurado. Entre em contato com a contabilidade.", "danger")
        return redirect(url_for("login"))
    rels = Relatorio.query.filter_by(posto_id=posto.id)\
                          .order_by(Relatorio.gerado_em.desc()).limit(20).all()
    return render_template("posto/dashboard.html",
                           posto=posto, relatorios=rels, usuario=current_user)


@app.route("/posto/conferir")
@requer_perfil(PERFIL_POSTO)
def posto_conferir():
    """Tela de upload do SPED para o posto."""
    posto = current_user.posto
    if not posto or not posto.licenca_ativa:
        flash("Licença inativa. Entre em contato com a Cleodon Contabilidade.", "danger")
        return redirect(url_for("posto_dashboard"))
    return render_template("posto/conferir.html", posto=posto, usuario=current_user)


@app.route("/posto/processar", methods=["POST"])
@requer_perfil(PERFIL_POSTO)
def posto_processar():
    """Processa o SPED enviado pelo posto e salva o relatório no histórico."""
    posto = current_user.posto
    if not posto or not posto.licenca_ativa:
        abort(403)

    arq_ant = request.files.get("ant")
    arq_atu = request.files.get("atu")
    tem_ant = bool(arq_ant and arq_ant.filename)

    if not arq_atu or not arq_atu.filename:
        flash("Selecione ao menos o SPED da competência atual.", "danger")
        return redirect(url_for("posto_conferir"))

    try:
        bytes_atu = arq_atu.read()
        d_atu = ler_sped_bytes(bytes_atu)

        if tem_ant:
            bytes_ant = arq_ant.read()
            d_ant   = ler_sped_bytes(bytes_ant)
            neg_abr = verificar_negativos_bytes(bytes_ant)
        else:
            d_ant   = {"info": {}, "tanques": {}, "bicos": {}}
            neg_abr = {"tanques": [], "bicos": []}

        conf_m  = confronto_mensal(d_ant, d_atu)
        d_mai   = confronto_diario(d_atu)
        neg_mai = verificar_negativos_bytes(bytes_atu)
        vc_mai  = verificar_versao_capacidade(d_atu)

        arq_dac = request.files.get("dac")
        conf_dac = None
        if arq_dac and arq_dac.filename:
            try:
                dac_dados = ler_dac(arq_dac.read(), arq_dac.filename)
                if dac_dados:
                    conf_dac = confrontar_dac_sped(dac_dados, d_atu)
            except Exception:
                pass

        # Gerar Excel
        cad_atu = verificar_cadastro_lmc(bytes_atu)

        wb = openpyxl.Workbook()
        ws1 = wb.active; ws1.title = "Resumo"
        aba_resumo(ws1, conf_m, d_mai, d_ant, d_atu, neg_abr, neg_mai, vc_mai, cad_atu)
        ws2 = wb.create_sheet("Confronto Meses"); aba_mensal(ws2, conf_m, d_ant, d_atu)
        ws3 = wb.create_sheet("Comparativo Diário"); aba_diario(ws3, d_mai)
        if conf_dac:
            ws4 = wb.create_sheet("DAC × SPED"); aba_dac(ws4, conf_dac, d_atu)
        ws5 = wb.create_sheet("DAC do SPED"); aba_dac_sped(ws5, d_atu, d_atu)
        ws6 = wb.create_sheet("Relatório ao Cliente")
        aba_relatorio_cliente(ws6, conf_m, d_mai, neg_abr, neg_mai,
                              vc_mai, conf_dac, d_ant, d_atu, cad_atu)

        ws7 = wb.create_sheet("Informação das Bombas", 1)
        aba_cadastro_lmc(ws7, cad_atu, d_atu)

        # Salvar arquivo na pasta do posto (isolado por CNPJ)
        pasta_posto = os.path.join(RELATORIOS_DIR, posto.cnpj)
        os.makedirs(pasta_posto, exist_ok=True)
        iu = d_atu.get("info", {})
        comp_atu = f"{iu.get('dt_ini','')[2:4]}_{iu.get('dt_ini','')[4:]}" or "competencia"
        nome_arquivo = f"LMC_{posto.cnpj}_{comp_atu}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        caminho = os.path.join(pasta_posto, nome_arquivo)
        wb.save(caminho)

        # Calcular divergências para o histórico
        def fmt_comp2(dt):
            meses=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
            try: return f"{meses[int(dt[2:4])-1]}/{dt[4:]}"
            except: return dt or "—"

        n_div = (sum(1 for x in conf_m["tanques"] if x["status"]!="✅ OK") +
                 sum(1 for x in conf_m["bicos"]   if x["status"]!="✅ OK") +
                 len(neg_mai["tanques"]) + len(neg_mai["bicos"]) +
                 sum(1 for t in conf_m["tanques"]
                     if t.get("status_anp","") not in ("✅ DENTRO DO LIMITE","")))
        status_geral = "ok" if n_div == 0 else ("critico" if n_div >= 5 else "alerta")

        rel = Relatorio(
            posto_id=posto.id,
            competencia_ant=fmt_comp2(d_ant.get("info",{}).get("dt_fin","")) if tem_ant else None,
            competencia_atu=fmt_comp2(iu.get("dt_fin","")),
            gerado_por=current_user.email,
            total_divergencias=n_div,
            tem_dac=bool(conf_dac),
            status_geral=status_geral,
            arquivo_nome=nome_arquivo,
        )
        db.session.add(rel); db.session.commit()

        # Enviar o arquivo para download
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        nome_dl = f"Relatorio_LMC_{_nome_posto(d_atu)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=nome_dl)

    except Exception as e:
        flash(f"Erro ao processar: {str(e)}", "danger")
        return redirect(url_for("posto_conferir"))


@app.route("/relatorio/<int:rel_id>/download")
@fl_login_required
def download_relatorio(rel_id):
    """Download de relatório do histórico — verificando permissão."""
    rel = Relatorio.query.get_or_404(rel_id)
    posto = rel.posto

    # Verificar permissão
    if current_user.is_posto and current_user.posto_id != posto.id:
        abort(403)
    if current_user.is_contabilidade:
        if posto.contabilidade_id != current_user.contabilidade_id:
            abort(403)

    caminho = os.path.join(RELATORIOS_DIR, posto.cnpj, rel.arquivo_nome)
    if not os.path.exists(caminho):
        flash("Arquivo não encontrado.", "danger")
        return redirect(url_for("dashboard"))

    return send_file(caminho, as_attachment=True,
                     download_name=rel.arquivo_nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═════════════════════════════════════════════════════════════════════════════
# PERFIL / ALTERAR SENHA
# ═════════════════════════════════════════════════════════════════════════════
@app.route("/perfil", methods=["GET","POST"])
@fl_login_required
def perfil():
    if request.method == "POST":
        senha_atual = request.form.get("senha_atual","")
        senha_nova  = request.form.get("senha_nova","")
        confirmacao = request.form.get("confirmacao","")
        if not current_user.check_senha(senha_atual):
            flash("Senha atual incorreta.", "danger")
        elif senha_nova != confirmacao:
            flash("Nova senha e confirmação não coincidem.", "danger")
        elif len(senha_nova) < 6:
            flash("A nova senha deve ter ao menos 6 caracteres.", "danger")
        else:
            current_user.set_senha(senha_nova)
            db.session.commit()
            flash("Senha alterada com sucesso!", "success")
    return render_template("perfil.html", usuario=current_user)


def aba_cadastro_lmc(ws, cad_atu, info_atu):
    """Aba de verificação de cadastro — tanques, bicos, bombas, aferições e lacres."""
    ws.sheet_view.showGridLines = False
    N = 5

    for l, w in zip("ABCDE", [18, 22, 22, 18, 18]):
        ws.column_dimensions[l].width = w

    iu = info_atu.get("info", {})

    def fmt_comp(dt):
        meses=["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
        try: return f"{meses[int(dt[2:4])-1]}/{dt[4:]}"
        except: return dt or "—"

    comp = fmt_comp(iu.get("dt_fin", ""))

    def aplic_brd(r, c1, c2):
        for col in range(c1, c2+1):
            ws.cell(row=r, column=col).border = _brd()

    def tit_sec(r, texto, ok):
        """Linha de seção: ✅ ou ❌ + título."""
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N)
        c = ws.cell(row=r, column=1, value=texto)
        bg = C_VERDE_BG if ok else C_VERM_BG
        fg = C_VERDE_FG if ok else C_VERM_FG
        c.font = Font(name="Arial", size=10, bold=True, color=fg)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        aplic_brd(r, 1, N)

    def cab(r, headers):
        col = 1
        for h in headers:
            c = ws.cell(row=r, column=col, value=h)
            c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", start_color="831040")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _brd()
            col += 1
        ws.row_dimensions[r].height = 18

    def lin(r, valores):
        for col, v in enumerate(valores, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Arial", size=9, color="1a2340")
            c.alignment = Alignment(horizontal="left" if col==1 else "center",
                                    vertical="center")
            c.border = _brd()
        ws.row_dimensions[r].height = 15

    row = 1

    # Título
    _titulo(ws, row, f"VERIFICAÇÃO DE CADASTRO LMC  –  {comp}", N, sz=12)
    ws.row_dimensions[row].height = 28; row += 1

    # Info
    for label, valor in [
        ("Empresa:", iu.get("razao", "")),
        ("CNPJ:", iu.get("cnpj", "")),
        ("Competência:", comp),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = Font(name="Arial", size=9, bold=True, color="555555")
        c1.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=N)
        c2 = ws.cell(row=row, column=3, value=valor)
        c2.font = Font(name="Arial", size=9, bold=(label=="Empresa:"), color="1a2340")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 15; row += 1
    row += 1

    # ── 1. TANQUES (1310) ─────────────────────────────────────────────────────
    sem_cap = cad_atu["tanques_sem_cap"]
    ok_t = cad_atu["tanques_ok"] and not sem_cap
    tit_sec(row,
        f"{'✅' if ok_t else '❌'}  TANQUES (Reg. 1310)  —  "
        f"{cad_atu['n_tanques']} tanque(s) declarado(s)"
        + (f"  |  {len(sem_cap)} sem capacidade" if sem_cap else ""),
        ok_t); row += 1

    if cad_atu["tanques_ok"]:
        cab(row, ["Tanque", "Capacidade (L)", "Status"]); row += 1
        for t in cad_atu["tanques"]:
            cap = cad_atu.get("tanques", {})
            # pegar capacidade diretamente do dict interno
            pass
        # Refazer: verificar_cadastro_lmc retorna lista; precisamos do dict completo
        # Vamos usar a info já processada — tanques é lista de IDs
        for t in cad_atu["tanques"]:
            lin(row, [f"Tanque {t}", "—", "⚠️ Sem dado de capacidade" if t in sem_cap else "✅ OK"])
            row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value="❌  Nenhum tanque declarado no registro 1310.")
        c.font = Font(name="Arial", size=9, color=C_VERM_FG)
        c.fill = PatternFill("solid", start_color=C_VERM_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        aplic_brd(row, 1, N); row += 1
    row += 1

    # ── 2. BICOS (1320) ───────────────────────────────────────────────────────
    ok_b = cad_atu["bicos_ok"]
    tit_sec(row,
        f"{'✅' if ok_b else '❌'}  BICOS (Reg. 1320)  —  "
        f"{cad_atu['n_bicos']} bico(s) declarado(s)", ok_b); row += 1

    if ok_b:
        cab(row, ["Bicos declarados"]); row += 1
        bicos_str = "  |  ".join([f"Bico {b}" for b in cad_atu["bicos"]])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value=bicos_str)
        c.font = Font(name="Arial", size=9, color="1a2340")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = _brd()
        ws.row_dimensions[row].height = 20; row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value="❌  Nenhum bico declarado no registro 1320.")
        c.font = Font(name="Arial", size=9, color=C_VERM_FG)
        c.fill = PatternFill("solid", start_color=C_VERM_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        aplic_brd(row, 1, N); row += 1
    row += 1

    # ── 3. BOMBAS (1350) ──────────────────────────────────────────────────────
    ok_bm = cad_atu["bombas_ok"]
    tit_sec(row,
        f"{'✅' if ok_bm else '❌'}  BOMBAS / MEDIDORES (Reg. 1350)  —  "
        f"{cad_atu['n_bombas']} bomba(s) cadastrada(s)", ok_bm); row += 1

    if ok_bm:
        cab(row, ["Série", "Fabricante", "Modelo", "Status"]); row += 1
        for b in cad_atu["bombas"]:
            ws.cell(row=row, column=1, value=b["serie"]).font = Font(name="Arial", size=9)
            ws.cell(row=row, column=2, value=b["fabricante"]).font = Font(name="Arial", size=9)
            ws.cell(row=row, column=3, value=b["modelo"]).font = Font(name="Arial", size=9)
            c_st = ws.cell(row=row, column=4, value="✅ OK")
            c_st.font = Font(name="Arial", size=9, color=C_VERDE_FG, bold=True)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=N)
            for col in range(1, N+1):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=row, column=col).border = _brd()
            ws.row_dimensions[row].height = 15; row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value="❌  Nenhuma bomba cadastrada no registro 1350.")
        c.font = Font(name="Arial", size=9, color=C_VERM_FG)
        c.fill = PatternFill("solid", start_color=C_VERM_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        aplic_brd(row, 1, N); row += 1
    row += 1

    # ── 4. LACRES (1360) ──────────────────────────────────────────────────────
    ok_af = cad_atu["afericoes_ok"]
    tit_sec(row,
        f"{'✅' if ok_af else '❌'}  LACRES (Reg. 1360)  —  "
        f"{cad_atu['n_afericoes']} lacre(s) declarado(s)", ok_af); row += 1

    if ok_af:
        cab(row, ["Nº Lacre", "Data Instalação", "Status"]); row += 1
        for af in cad_atu["afericoes"]:
            data_fmt = af["data"]
            if len(data_fmt) == 8:
                data_fmt = f"{data_fmt[:2]}/{data_fmt[2:4]}/{data_fmt[4:]}"
            lin(row, [af["numero"], data_fmt, "✅ OK"]); row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value="❌  Nenhum lacre declarado no registro 1360.")
        c.font = Font(name="Arial", size=9, color=C_VERM_FG)
        c.fill = PatternFill("solid", start_color=C_VERM_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        aplic_brd(row, 1, N); row += 1
    row += 1

    # ── 5. BICOS POR TANQUE/PRODUTO (1370) ───────────────────────────────────
    ok_lc = cad_atu["lacres_ok"]
    tit_sec(row,
        f"{'✅' if ok_lc else '❌'}  BICOS POR TANQUE E PRODUTO (Reg. 1370)  —  "
        f"{cad_atu['n_lacres']} bico(s) declarado(s)", ok_lc); row += 1

    if ok_lc:
        cab(row, ["Nº Bico", "Produto", "Tanque"]); row += 1
        for bl in sorted(cad_atu["bicos_lacres"],
                         key=lambda x: (x["tanque"], int(x["bico"]) if x["bico"].isdigit() else 0)):
            prod_nome = PRODUTOS_COMBUSTIVEL.get(bl["produto"], bl["produto"])
            lin(row, [f"Bico {bl['bico']}", prod_nome, f"Tanque {bl['tanque']}"]); row += 1
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1, value="❌  Nenhum vínculo bico-produto declarado no registro 1370.")
        c.font = Font(name="Arial", size=9, color=C_VERM_FG)
        c.fill = PatternFill("solid", start_color=C_VERM_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        aplic_brd(row, 1, N); row += 1


# ═════════════════════════════════════════════════════════════════════════════
# ACOMPANHAMENTO DE RECEBIMENTO DE COMBUSTÍVEIS
# ═════════════════════════════════════════════════════════════════════════════
from recebimento import ler_notas_xls, ler_estoque_dac, confrontar_recebimento

@app.route("/recebimento", methods=["GET", "POST"])
@fl_login_required
def recebimento():
    if current_user.is_posto:
        return redirect(url_for("posto_dashboard"))

    resultado = None
    erro = None

    if request.method == "POST":
        arq_notas = request.files.get("notas")
        arq_dac   = request.files.get("dac")

        if not arq_notas or not arq_dac:
            erro = "Selecione os dois arquivos (Excel de notas e PDF do DAC)."
        else:
            try:
                bytes_notas = arq_notas.read()
                bytes_dac   = arq_dac.read()

                notas_data, err1 = ler_notas_xls(bytes_notas, arq_notas.filename)
                dac_data,   err2 = ler_estoque_dac(bytes_dac, arq_dac.filename)

                if err1: erro = f"Erro ao ler notas: {err1}"
                elif err2: erro = f"Erro ao ler DAC: {err2}"
                else:
                    linhas = confrontar_recebimento(notas_data, dac_data)

                    # Opção de download Excel
                    if request.form.get("acao") == "download":
                        import openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        from openpyxl.utils import get_column_letter

                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "Recebimento"
                        ws.sheet_view.showGridLines = False

                        def brd():
                            s = Side(style='thin', color='D0D8E4')
                            return Border(left=s, right=s, top=s, bottom=s)

                        # Título linha 1: Nome da empresa + Competência
                        empresa_titulo = dac_data.get('empresa','') or ''
                        comp_titulo    = dac_data.get('competencia','') or ''
                        ws.merge_cells('A1:E1')
                        c = ws.cell(row=1, column=1,
                            value=f"{empresa_titulo}  {comp_titulo}".strip())
                        c.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
                        c.fill = PatternFill("solid", start_color="CC0066")
                        c.alignment = Alignment(horizontal="center", vertical="center")
                        for col in range(1,6): ws.cell(row=1,column=col).border = brd()
                        ws.row_dimensions[1].height = 26

                        # Cabeçalho tabela
                        headers = ["PRODUTO", "NOTAS (L)", "DAC (L)", "PERDA (L)", "GANHO (L)"]
                        for col, h in enumerate(headers, 1):
                            c = ws.cell(row=5, column=col, value=h)
                            c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                            c.fill = PatternFill("solid", start_color="990066")
                            c.alignment = Alignment(horizontal="center", vertical="center")
                            c.border = brd()
                        ws.row_dimensions[5].height = 20

                        # Dados
                        for i, l in enumerate(linhas, 6):
                            def cel(col, val, fmt=None):
                                c = ws.cell(row=i, column=col, value=val)
                                c.font = Font(name="Arial", size=9, color="1a2340")
                                c.alignment = Alignment(
                                    horizontal="left" if col==1 else "center",
                                    vertical="center")
                                c.border = brd()
                                if fmt: c.number_format = fmt
                            cel(1, l['produto'])
                            cel(2, l['qtd_notas'] or 0, '#,##0.00')
                            cel(3, l['rec_dac']   or 0, '#,##0.00')
                            cel(4, l['perda']     or 0, '#,##0.00')
                            cel(5, l['ganho']     or 0, '#,##0.00')
                            ws.row_dimensions[i].height = 15

                        for col, w in zip("ABCDE", [28,16,16,16,16]):
                            ws.column_dimensions[col].width = w

                        # Notas detalhadas
                        row_n = len(linhas) + 4  # 1 título + 1 cabeçalho + n dados + 1 espaço
                        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=4)
                        c = ws.cell(row=row_n, column=1, value="NOTAS FISCAIS DE ENTRADA")
                        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                        c.fill = PatternFill("solid", start_color="990066")
                        c.alignment = Alignment(horizontal="left", vertical="center")
                        for col in range(1,5): ws.cell(row=row_n,column=col).border = brd()
                        ws.row_dimensions[row_n].height = 20
                        row_n += 1

                        for col, h in enumerate(["NF", "Data", "Produto", "Qtd. (L)"], 1):
                            c = ws.cell(row=row_n, column=col, value=h)
                            c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                            c.fill = PatternFill("solid", start_color="831040")
                            c.alignment = Alignment(horizontal="center", vertical="center")
                            c.border = brd()
                        row_n += 1

                        for n in notas_data['notas']:
                            ws.cell(row=row_n, column=1, value=n['nf']).border = brd()
                            ws.cell(row=row_n, column=2, value=n['data'].strftime('%d/%m/%Y') if n['data'] else '').border = brd()
                            ws.cell(row=row_n, column=3, value=n['produto']).border = brd()
                            c_q = ws.cell(row=row_n, column=4, value=n['quantidade'])
                            c_q.number_format = '#,##0.000'; c_q.border = brd()
                            for col in range(1, 5):
                                ws.cell(row=row_n, column=col).font = Font(name="Arial", size=9)
                                ws.cell(row=row_n, column=col).alignment = Alignment(
                                    horizontal="left" if col in (1,2,3) else "center",
                                    vertical="center")
                            ws.row_dimensions[row_n].height = 14
                            row_n += 1

                        buf = io.BytesIO()
                        wb.save(buf); buf.seek(0)
                        nome_dl = f"Recebimento_{dac_data.get('competencia','').replace('/','_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        return send_file(buf,
                            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            as_attachment=True, download_name=nome_dl)

                    resultado = {
                        'linhas':      linhas,
                        'notas':       notas_data['notas'],
                        'totais':      notas_data['totais'],
                        'competencia': dac_data.get('competencia', ''),
                        'empresa':     dac_data.get('empresa', ''),
                    }
            except Exception as e:
                erro = f"Erro ao processar: {str(e)}"

    return render_template("recebimento.html",
                           resultado=resultado, erro=erro, usuario=current_user)
