"""
Módulo de leitura do DAC — sem dependência de API externa.
Suporta:
  - Excel (.xlsx, .xls): leitura direta via openpyxl/xlrd
  - PDF com texto extraível: parse automático (formato AutoSystem PRO e similares)
  - PDF escaneado / imagem: retorna None (sistema gera sem confronto DAC)
"""
import io, re, json

def _fl(v):
    """Converte para float. Aceita number nativo (int/float, ex: do openpyxl)
    OU string no formato brasileiro: '1.234,56' -> 1234.56"""
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v)
    v = str(v).strip().replace(".", "").replace(",", ".")
    try: return float(v)
    except: return None

def _nid(v):
    """Normaliza ID: '01' → '1', '003' → '3'"""
    try: return str(int(str(v).strip()))
    except: return str(v).strip()


# ── EXCEL ─────────────────────────────────────────────────────────────────────
def _ler_excel(arquivo_bytes, filename):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), data_only=True)
    # Concatenar todas as abas em texto tabular
    linhas = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() if v is not None else "" for v in row]
            linha = " | ".join(vals).strip(" |")
            if linha:
                linhas.append(linha)
    texto = "\n".join(linhas)
    return _parse_texto(texto)


# ── PDF ───────────────────────────────────────────────────────────────────────
def _ler_pdf(arquivo_bytes):
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(arquivo_bytes))
        texto = ""
        for page in reader.pages:
            texto += (page.extract_text() or "") + "\n"
        texto = texto.strip()
        if len(texto) < 50:
            return None  # PDF escaneado/imagem — sem texto extraível

        # Tentar parser "Relatório DAC Referente ao Mês" (variante legível: Tulemon, Gol)
        if re.search(r'RELAT[ÓO]RIO\s*DAC\s*REFERENTE\s*AO\s*M[ÊE]S', texto, re.I):
            resultado = _parse_relatorio_dac_mes(texto)
            if resultado["tanques"] or resultado["bicos"]:
                return resultado

        # Tentar parser "Resumo do Livro de Movimentação (R-LMC)" - multi-página por tanque
        if re.search(r'RESUMO\s*DO\s*LIVRO\s*DE\s*MOVIMENTA[ÇC][ÃA]O', texto, re.I):
            resultado = _parse_resumo_lmc(texto)
            if resultado["tanques"] or resultado["bicos"]:
                return resultado

        # Tentar parser "Resumo DAC" primeiro (formato com 1 linha por bico/tanque)
        if re.search(r'Resumo\s*DAC', texto, re.I):
            resultado = _parse_resumo_dac(texto)
            if resultado["tanques"] or resultado["bicos"]:
                return resultado

        # Tentar parser "Declaração de Atividades" - duas variantes:
        # 1) linear (1 linha por bico/tanque, ex: GrupoZL)
        # 2) verticalizada (campos em linhas separadas, ex: outro sistema)
        if re.search(r'DECLARA[ÇC][ÃA]O\s*DE\s*ATIVIDADES', texto, re.I):
            resultado = _parse_declaracao_linear(texto)
            if resultado["tanques"] or resultado["bicos"]:
                return resultado
            resultado = _parse_declaracao_atividades(texto)
            if resultado["tanques"] or resultado["bicos"]:
                return resultado

        # Fallback: parser AutoSystem PRO (multi-linha por campo)
        resultado = _parse_texto(texto)
        if resultado["tanques"] or resultado["bicos"]:
            return resultado

        return None
    except Exception:
        return None


# ── PARSER FORMATO "RESUMO DAC" (uma linha por bico/tanque) ──────────────────
def _parse_resumo_dac(texto):
    """
    Parser para o formato 'Resumo DAC' onde cada bico/tanque está em UMA linha:
      Bomba Bico Combustível Enc.Inicial Enc.Final Vol.semInt. Aferição Vol.comInt.
      Tanque Combustível Est.Abertura Est.Fechamento Vol.Recebido Faltas/Sobras
    """
    tanques = []
    bicos   = []
    competencia = ""

    # Competência
    m = re.search(r'[Pp]er[íi]odo\s+de\s+apura[çc][ãa]o[:\s]+(\d{2}/\d{2}/\d{4})\s+at[ée]\s+(\d{2}/\d{2}/\d{4})', texto)
    if m:
        dt_fin = m.group(2)
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try:
            mes = int(dt_fin[3:5]); ano = dt_fin[6:]
            competencia = f"{meses[mes-1]}/{ano}"
        except:
            competencia = dt_fin

    num = r'-?[\d\.]+,\d+'  # número brasileiro: 1.234,567 ou -123,456

    # ── BICOS: "001 001 GASOLINA COMUM 1.655.772,200 1.681.591,360 0,000 25.819,160 1,211"
    # Padrão: bomba(3) bico(3) PRODUTO(palavras) num num num num num
    bico_pat = re.compile(
        r'^(\d{2,3})\s+(\d{2,3})\s+([A-ZÀ-Ú0-9 ]+?)\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')(?:\s+(' + num + r'))?\s*$'
    )
    for linha in texto.splitlines():
        l = linha.strip()
        m = bico_pat.match(l)
        if m:
            bico_id = _nid(m.group(2))
            enc_ini = _fl(m.group(4))
            enc_fin = _fl(m.group(5))
            if enc_ini is not None and enc_fin is not None and enc_fin >= enc_ini:
                # Evita capturar linhas de bombas (modelo/série) por engano:
                # só aceita se já vimos ENCERRANTE crescente (bico real)
                ids_existentes = [b['id'] for b in bicos]
                if bico_id not in ids_existentes:
                    bicos.append({
                        "id": bico_id,
                        "encerrante_inicial": enc_ini,
                        "encerrante_final":   enc_fin,
                    })

    # ── TANQUES: "001 GASOLINA COMUM 10.840,862 5.618,234 149.000,000 -31.095,102"
    # Padrão real extraído: "001 10.840,862GASOLINA COMUM 5.618,234 149.000,000 -31.095,102"
    # (PyPDF às vezes gruda o número com o texto seguinte sem espaço)
    tanque_pat = re.compile(
        r'^(\d{3})\s+(' + num + r')\s*([A-ZÀ-Ú][A-ZÀ-Ú0-9 ]*?)\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')\s*$'
    )
    em_estoque = False
    for linha in texto.splitlines():
        l = linha.strip()
        if re.search(r'Estoque\s*F[íi]sico', l, re.I):
            em_estoque = True; continue
        if re.search(r'^Resumo\s*DAC\s*-', l, re.I):
            em_estoque = False; continue
        if not em_estoque:
            continue
        m = tanque_pat.match(l)
        if m:
            tid     = _nid(m.group(1))
            ini     = _fl(m.group(2))
            produto = m.group(3).strip()
            fin     = _fl(m.group(4))
            ids_existentes = [t['id'] for t in tanques]
            if tid not in ids_existentes:
                tanques.append({
                    "id": tid,
                    "produto": produto,
                    "estoque_inicial": ini,
                    "estoque_final":   fin,
                })

    return {"competencia": competencia, "tanques": tanques, "bicos": bicos}


# ── PARSER FORMATO "RELATÓRIO DE DECLARAÇÃO DE ATIVIDADES" (campos verticais) ─
def _parse_declaracao_atividades(texto):
    """
    Parser para o formato 'RELATÓRIO DE DECLARAÇÃO DE ATIVIDADES DO CONTRIBUINTE'.
    Cada campo de uma linha de tabela vem em uma linha separada do PDF
    (extração verticalizada), em blocos de tamanho fixo:

      BICO (10 linhas/bico): Lacre, Aferição, ComInterv, Bomba,
                              EncFinal, EncInicial, Combustível, Tanque, Bico, SemInterv

      TANQUE (7 linhas/tanque): PerdaSobra, EstFechamento, Vendas,
                                 Recebimentos, EstAbertura, Item, Tanque
    """
    tanques = []
    bicos   = []
    competencia = ""

    datas = re.findall(r"(\d{2}/\d{2}/\d{4})", texto[:600])
    if len(datas) >= 2:
        dt_fin = datas[1]
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try:
            competencia = f"{meses[int(dt_fin[3:5])-1]}/{dt_fin[6:]}"
        except:
            competencia = dt_fin

    linhas = [l.strip() for l in texto.splitlines()]
    num_re = re.compile(r'^-?[\d\.]+,\d+$')
    int_re = re.compile(r'^\d+$')

    # ── Seção BICO ────────────────────────────────────────────────────────────
    idx_bico_start = next((i for i,l in enumerate(linhas) if 'MOVIMENTA' in l.upper() and 'BICO' in l.upper()), None)
    idx_bico_end   = next((i for i,l in enumerate(linhas) if 'MOVIMENTA' in l.upper() and 'TANQUE' in l.upper()), None)

    if idx_bico_start is not None and idx_bico_end is not None:
        bloco = linhas[idx_bico_start:idx_bico_end]
        # Pular cabeçalho (achar primeira linha que é um inteiro pequeno = Lacre)
        # Avançamos até encontrar o primeiro grupo de 10 valores de dados
        i = 0
        # Localizar onde os dados começam: primeiro token que é número inteiro
        # seguido pelos padrões esperados. Vamos escanear em janelas de 10.
        dados = [l for l in bloco if l]  # remove vazios
        # Remover cabeçalho textual (tudo antes do primeiro valor numérico isolado
        # que aparece logo seguido por um número decimal)
        start_idx = None
        for j in range(len(dados)-1):
            if int_re.match(dados[j]) and num_re.match(dados[j+1]):
                start_idx = j
                break
        if start_idx is not None:
            grupo = dados[start_idx:]
            # Processar em blocos de 10
            for k in range(0, len(grupo)-9, 10):
                bloco10 = grupo[k:k+10]
                try:
                    lacre      = bloco10[0]
                    aferic     = bloco10[1]
                    cominterv  = bloco10[2]
                    bomba      = bloco10[3]
                    enc_final  = _fl(bloco10[4])
                    enc_inicial= _fl(bloco10[5])
                    combustivel= bloco10[6]
                    tanque_id  = bloco10[7]
                    bico_id    = bloco10[8]
                    seminterv  = bloco10[9]
                    if not int_re.match(bico_id):
                        continue
                    if enc_inicial is not None and enc_final is not None:
                        bicos.append({
                            "id": _nid(bico_id),
                            "encerrante_inicial": enc_inicial,
                            "encerrante_final":   enc_final,
                        })
                except (IndexError, ValueError):
                    continue

    # ── Seção TANQUE ──────────────────────────────────────────────────────────
    idx_tanque_start = idx_bico_end
    idx_tanque_end = next((i for i,l in enumerate(linhas) if 'MOVIMENTA' in l.upper() and 'CFOP' in l.upper()), len(linhas))

    if idx_tanque_start is not None:
        bloco = linhas[idx_tanque_start:idx_tanque_end]
        dados = [l for l in bloco if l]
        # Localizar início dos dados: primeiro valor decimal (pode ser negativo)
        # seguido pelo padrão de 7 campos terminando em um inteiro pequeno (tanque)
        start_idx = None
        for j in range(len(dados)-6):
            if num_re.match(dados[j]):
                # verificar se 6 posições à frente é um inteiro pequeno (tanque 1-99)
                cand = dados[j+6]
                if int_re.match(cand) and int(cand) <= 99:
                    start_idx = j
                    break
        if start_idx is not None:
            grupo = dados[start_idx:]
            for k in range(0, len(grupo)-6, 7):
                bloco7 = grupo[k:k+7]
                try:
                    perda_sobra = _fl(bloco7[0])
                    est_fech    = _fl(bloco7[1])
                    vendas      = _fl(bloco7[2])
                    recebim     = _fl(bloco7[3])
                    est_abert   = _fl(bloco7[4])
                    item        = bloco7[5]
                    tanque_id   = bloco7[6]
                    if not int_re.match(tanque_id):
                        continue
                    if est_abert is not None and est_fech is not None:
                        tanques.append({
                            "id": _nid(tanque_id),
                            "produto": item,
                            "estoque_inicial": est_abert,
                            "estoque_final":   est_fech,
                        })
                except (IndexError, ValueError):
                    continue

    return {"competencia": competencia, "tanques": tanques, "bicos": bicos}


# ── PARSER FORMATO "DECLARAÇÃO DE ATIVIDADES - LINEAR" (1 linha por item) ────
def _parse_declaracao_linear(texto):
    """
    Parser para variante do formato 'Declaração de Atividades do Contribuinte'
    onde cada bico/tanque está em UMA linha (ex: GrupoZL):
      Bomba Bico Tanque Combustível EncIni EncFim SemInterv ComInterv Aferição
      Tanque Item Capacidade EstAbert Recebim Vendas EstFech Perda/Sobra Var%
    """
    tanques = []
    bicos   = []
    competencia = ""

    m = re.search(r'(\d{2}/\d{2}/\d{4})\s+at[ée]\s+(\d{2}/\d{2}/\d{4})', texto)
    if m:
        dt_fin = m.group(2)
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try: competencia = f"{meses[int(dt_fin[3:5])-1]}/{dt_fin[6:]}"
        except: competencia = dt_fin

    num = r'-?[\d\.]+,\d+'

    # BICOS: "1 1 3 DIESEL S-10 ADITIVADO 79.717,887 82.342,868 2.624,981 0,000 0,000"
    bico_pat = re.compile(
        r'^(\d{1,2})\s+(\d{1,2})\s+(\d{1,2})\s+([A-ZÀ-Úa-zà-ú0-9\- ]+?)\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')\s*$'
    )
    ids_vistos = set()
    for linha in texto.splitlines():
        l = linha.strip()
        m = bico_pat.match(l)
        if m:
            bico_id = _nid(m.group(2))
            enc_ini = _fl(m.group(5))
            enc_fin = _fl(m.group(6))
            if bico_id not in ids_vistos and enc_ini is not None and enc_fin is not None:
                bicos.append({"id": bico_id, "encerrante_inicial": enc_ini, "encerrante_final": enc_fin})
                ids_vistos.add(bico_id)

    # TANQUES: "1 DIESEL S-10 ADITIVADO 3.631,000 49.700,000 53.282,813 123,000 74,813 1,96220.000,00"
    # Ordem: id, produto..., EstAbert, Recebim, Vendas, EstFech, Perda/Sobra, Var%(+Capacidade colada no fim)
    tanque_pat = re.compile(
        r'^(\d{1,2})\s+([A-ZÀ-Úa-zà-ú0-9\- ]+?)\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')\s+(-?' + num + r')\s+(-?' + num + r')\s+(-?[\d\.]+,\d+)'
    )
    ids_t = set()
    em_tanques = False
    for linha in texto.splitlines():
        l = linha.strip()
        if re.search(r'MOVIMENTA[ÇC][ÃA]O\s*POR\s*TANQUE', l, re.I):
            em_tanques = True; continue
        if re.search(r'MOVIMENTA[ÇC][ÃA]O\s*POR\s*CFOP', l, re.I):
            em_tanques = False; continue
        if not em_tanques: continue
        m = tanque_pat.match(l)
        if m:
            tid = _nid(m.group(1))
            if tid in ids_t: continue
            produto   = m.group(2).strip()
            est_abert = _fl(m.group(3))
            est_fech  = _fl(m.group(6))
            if est_abert is not None and est_fech is not None:
                tanques.append({"id": tid, "produto": produto,
                                "estoque_inicial": est_abert, "estoque_final": est_fech})
                ids_t.add(tid)

    return {"competencia": competencia, "tanques": tanques, "bicos": bicos}


# ── PARSER FORMATO "RESUMO DO LIVRO DE MOVIMENTAÇÃO (R-LMC)" ─────────────────
def _parse_resumo_lmc(texto):
    """
    Parser para o formato 'RESUMO DO LIVRO DE MOVIMENTAÇÃO DE COMBUSTÍVEIS (R-LMC)'.
    Multi-página: uma página (ou bloco) por tanque, contendo:
      - 3.1) Estoque de Abertura (do tanque, primeiro dia do período)
      - 7) Estoque de Fechamento (do tanque, último dia do período)
      - 5) Volume Vendido por Bico: TQ, Bico, Fechamento, Abertura, Aferições, VendasBico
      - TANQUE - Nº: X (identifica o tanque)
    Considera apenas abertura/fechamento (ignora notas fiscais e detalhes diários).
    """
    tanques = []
    bicos   = []
    competencia = ""

    num = r'-?[\d\.]+,\d+'

    m = re.search(r'D\.\s*Final\s*\n?\s*(\d{2}/\d{2}/\d{4})', texto)
    if not m:
        m = re.search(r'(\d{2}/\d{2}/\d{4})\s*\n?\s*D\.\s*Final', texto)
    datas = re.findall(r'(\d{2}/\d{2}/\d{4})', texto[:400])
    if len(datas) >= 2:
        dt_fin = datas[1]
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try: competencia = f"{meses[int(dt_fin[3:5])-1]}/{dt_fin[6:]}"
        except: competencia = dt_fin

    # Identificar blocos por "TANQUE - Nº: X" (cada bloco = 1 tanque consolidado)
    # O nome do produto fica ANTES de "1) Produto", então cortamos incluindo
    # o texto residual do bloco anterior (overlap) para capturar o produto certo.
    marcadores = [m.start() for m in re.finditer(r'1\)\s*Produto', texto)]
    blocos = []
    for i, pos in enumerate(marcadores):
        fim = marcadores[i+1] if i+1 < len(marcadores) else len(texto)
        blocos.append(texto[pos:fim])
    # Produto de cada bloco = últimas linhas não vazias do texto ANTES do marcador
    produtos_antes = []
    for i, pos in enumerate(marcadores):
        inicio_busca = marcadores[i-1] if i>0 else 0
        trecho_antes = texto[inicio_busca:pos]
        m_prod = re.search(r'([A-ZÀ-Ú][A-ZÀ-Úa-zà-ú0-9 \-]{3,})\s*\n\s*2\)\s*D\.\s*Inicial', trecho_antes)
        produtos_antes.append(m_prod.group(1).strip() if m_prod else "")

    bicos_vistos = set()

    for idx_bloco, bloco in enumerate(blocos):
        produto = produtos_antes[idx_bloco] if idx_bloco < len(produtos_antes) else ""

        # Número do tanque: "TANQUE - Nº: 5"
        m_tq = re.search(r'TANQUE\s*-\s*N[ºo]:\s*(\d+)', bloco)
        if not m_tq:
            continue
        tanque_id = _nid(m_tq.group(1))

        # 3.1) Estoque de Abertura
        m_ea = re.search(r'3\.1\)\s*Estoque\s*de\s*Abertura\s*\n?\s*(' + num + r')', bloco)
        est_abert = _fl(m_ea.group(1)) if m_ea else None

        # 7) Estoque de Fechamento (9.1)
        m_ef = re.search(r'7\)\s*Estoque\s*de\s*Fechamento\s*\(9\.1\)\s*\n?\s*(' + num + r')', bloco)
        est_fech = _fl(m_ef.group(1)) if m_ef else None

        if est_abert is not None and est_fech is not None:
            ids_existentes = [t['id'] for t in tanques]
            if tanque_id not in ids_existentes:
                tanques.append({
                    "id": tanque_id, "produto": produto,
                    "estoque_inicial": est_abert, "estoque_final": est_fech,
                })

        # ── Bicos do bloco: "5\n1\n1.472.864,980\n1.457.391,490\n0,000\n15.473,490"
        # Ordem: TQ, Bico, Fechamento, Abertura, Aferições, VendasBico
        idx_v = bloco.find('5) Volume Vendido')
        idx_p = bloco.find('8) Perdas')
        if idx_v == -1: continue
        sub = bloco[idx_v: idx_p if idx_p > idx_v else idx_v+2000]
        linhas_sub = [l.strip() for l in sub.splitlines() if l.strip()]

        int_re = re.compile(r'^\d{1,3}$')
        num_re = re.compile(r'^' + num + r'$')

        # Localizar início dos dados (pula cabeçalho textual)
        start = None
        for j in range(len(linhas_sub)-5):
            if (int_re.match(linhas_sub[j]) and int_re.match(linhas_sub[j+1])
                and num_re.match(linhas_sub[j+2]) and num_re.match(linhas_sub[j+3])):
                start = j
                break
        if start is None:
            continue

        k = start
        while k+5 < len(linhas_sub):
            tq_b   = linhas_sub[k]
            bico_b = linhas_sub[k+1]
            fech_b = linhas_sub[k+2]
            aber_b = linhas_sub[k+3]
            afer_b = linhas_sub[k+4]
            vend_b = linhas_sub[k+5]
            if not (int_re.match(tq_b) and int_re.match(bico_b) and num_re.match(fech_b) and num_re.match(aber_b)):
                break
            bico_id = _nid(bico_b)
            enc_fin = _fl(fech_b)
            enc_ini = _fl(aber_b)
            if bico_id not in bicos_vistos and enc_ini is not None and enc_fin is not None:
                bicos.append({
                    "id": bico_id,
                    "encerrante_inicial": enc_ini,
                    "encerrante_final":   enc_fin,
                })
                bicos_vistos.add(bico_id)
            k += 6

    return {"competencia": competencia, "tanques": tanques, "bicos": bicos}


# ── PARSER FORMATO "RELATÓRIO DAC REFERENTE AO MÊS" (Tulemon/Gol legível) ────
def _parse_relatorio_dac_mes(texto):
    """
    Parser para variante legível do 'RELATÓRIO DAC REFERENTE AO MÊS: MM/AAAA'.

    BICOS (uma linha por bico, dentro de blocos "Combustível: ..."):
      "1 1.191.301,23 1.213.535,89 20,003 794316 22.234,66 0,00"
       id  abertura      fechamento   afer+TQ  num_serie  s/interv  c/interv

    TANQUES (uma linha por tanque):
      "1 ETANOL HIDRATADO COMUM 3.573,94 3.950,71 55.000,00 13,22ETA"
       id   produto              est_abert  est_fech   recebido  perda/sobra+sigla
    """
    tanques = []
    bicos   = []
    competencia = ""

    m = re.search(r'REFERENTE\s*AO\s*M[ÊE]S:\s*(\d{2})/(\d{4})', texto, re.I)
    if m:
        meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
        try: competencia = f"{meses[int(m.group(1))-1]}/{m.group(2)}"
        except: competencia = f"{m.group(1)}/{m.group(2)}"

    num = r'-?[\d\.]+,\d+'

    # ── BICOS ─────────────────────────────────────────────────────────────────
    # "1 1.191.301,23 1.213.535,89 20,003 794316 22.234,66 0,00"
    bico_pat = re.compile(
        r'^(\d{1,2})\s+(' + num + r')\s+(' + num + r')\s+' + num + r'(\d{1,2})\s+(\w+)\s+(' + num + r')\s+(' + num + r')\s*$'
    )
    ids_vistos = set()
    em_bicos = False
    for linha in texto.splitlines():
        l = linha.strip()
        if re.search(r'INFORMA[ÇC][ÕO]ES\s*MENSAIS\s*DOS\s*ENCERRANTES', l, re.I):
            em_bicos = True; continue
        if re.search(r'INFORMA[ÇC][ÕO]ES\s*MENSAIS\s*DOS\s*ESTOQUES', l, re.I):
            em_bicos = False; continue
        if not em_bicos: continue
        m2 = bico_pat.match(l)
        if m2:
            bico_id = _nid(m2.group(1))
            enc_ini = _fl(m2.group(2))
            enc_fin = _fl(m2.group(3))
            if bico_id not in ids_vistos and enc_ini is not None and enc_fin is not None:
                bicos.append({
                    "id": bico_id,
                    "encerrante_inicial": enc_ini,
                    "encerrante_final":   enc_fin,
                })
                ids_vistos.add(bico_id)

    # ── TANQUES ───────────────────────────────────────────────────────────────
    # "1 ETANOL HIDRATADO COMUM 3.573,94 3.950,71 55.000,00 13,22ETA"
    tanque_pat = re.compile(
        r'^(\d{1,2})\s+([A-ZÀ-Úa-zà-ú0-9 \-]+?)\s+(' + num + r')\s+(' + num + r')\s+(' + num + r')\s+(-?[\d\.]+,\d{2})[A-Z0-9]{2,4}\s*$'
    )
    ids_t = set()
    em_tanques = False
    for linha in texto.splitlines():
        l = linha.strip()
        if re.search(r'INFORMA[ÇC][ÕO]ES\s*MENSAIS\s*DOS\s*ESTOQUES', l, re.I):
            em_tanques = True; continue
        if re.search(r'^P[áa]gina', l, re.I):
            em_tanques = False; continue
        if not em_tanques: continue
        m3 = tanque_pat.match(l)
        if m3:
            tid = _nid(m3.group(1))
            if tid in ids_t: continue
            produto   = m3.group(2).strip()
            est_abert = _fl(m3.group(3))
            est_fech  = _fl(m3.group(4))
            if est_abert is not None and est_fech is not None:
                tanques.append({"id": tid, "produto": produto,
                                "estoque_inicial": est_abert, "estoque_final": est_fech})
                ids_t.add(tid)

    return {"competencia": competencia, "tanques": tanques, "bicos": bicos}


# ── PARSER UNIVERSAL ──────────────────────────────────────────────────────────
def _parse_texto(texto):
    """
    Parser para o formato AutoSystem PRO / Linx e similares.
    Extrai tanques e bicos da seção POSIÇÃO DOS TANQUES e BICO E ENCERRANTES.
    """
    tanques = []
    bicos   = []
    competencia = ""

    linhas = texto.splitlines()

    # Detectar período/competência
    for l in linhas:
        m = re.search(r'[Pp]er[íi]odo[:\s]+(\d{2}/\d{2}/\d{4})\s+a\s+(\d{2}/\d{2}/\d{4})', l)
        if m:
            dt_fin = m.group(2)  # "30/04/2026"
            meses = ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
            try:
                mes = int(dt_fin[3:5]); ano = dt_fin[6:]
                competencia = f"{meses[mes-1]}/{ano}"
            except:
                competencia = dt_fin
            break

    # ── Seção BICO E ENCERRANTES ─────────────────────────────────────────────
    # Formato: "01 GC\n91.387,190\n108.958,710\n..." (uma linha por campo)
    # Ou: "01 GC | 91.387,190 | 108.958,710"
    em_bicos = False
    bico_buffer = []  # acumula tokens de um bico

    for i, l in enumerate(linhas):
        l = l.strip()
        if re.search(r'BICO\s*E\s*ENCERRANTE', l, re.I):
            em_bicos = True; continue
        if re.search(r'POSI[ÇC][ÃA]O\s*DOS\s*(COMBUST|TANQUE)', l, re.I):
            em_bicos = False; continue
        if not em_bicos: continue
        if re.search(r'^(Bico|Inicial|Final|Litros|Afer|Venda|Pre[çc]|Desc|Valor|Total)', l, re.I):
            continue

        # Linha de bico: começa com número seguido de letras/dígitos (ex: "01 GC", "02 0GC")
        m = re.match(r'^(\d{1,2})\s+[A-Z0-9]{1,5}$', l)
        if m:
            bico_buffer = [m.group(1)]  # novo bico
            continue

        # Acumular valores numéricos para o bico atual
        if bico_buffer and re.match(r'^[\d.,]+$', l.replace(".", "").replace(",", "")):
            bico_buffer.append(l)
            if len(bico_buffer) >= 3:  # id + inicial + final
                bid = _nid(bico_buffer[0])
                enc_ini = _fl(bico_buffer[1])
                enc_fin = _fl(bico_buffer[2])
                if enc_ini is not None and enc_fin is not None:
                    bicos.append({
                        "id": bid,
                        "encerrante_inicial": enc_ini,
                        "encerrante_final":   enc_fin,
                    })
                bico_buffer = []

    # ── Seção POSIÇÃO DOS TANQUES ─────────────────────────────────────────────
    # Formato: "TANQUE 001 /// GAS\n7.980,83\n79.000,00\n...\n6.851,10\n..."
    em_tanques = False
    tanq_buffer = []

    for l in linhas:
        l = l.strip()
        if re.search(r'POSI[ÇC][ÃA]O\s*DOS\s*TANQUE', l, re.I):
            em_tanques = True; continue
        if re.search(r'^\*Entrada', l, re.I):
            em_tanques = False; continue
        if not em_tanques: continue
        if re.search(r'^(Produto|Início|Entrada|Venda|Afer|Final|Medi|Difer|Total)', l, re.I):
            continue

        # Linha de tanque: "TANQUE 001 /// GAS" ou "TANQUE001///GAS"
        m = re.match(r'TANQUE\s*(\d+)\s*[/\\|]+\s*(\w+)', l, re.I)
        if m:
            tanq_buffer = [m.group(1), m.group(2)]
            continue

        if tanq_buffer and re.match(r'^-?[\d.,]+$', l.replace(".", "").replace(",","")):
            tanq_buffer.append(l)
            # Precisamos de pelo menos: id, produto, inicio, entrada, venda, afer, final
            if len(tanq_buffer) >= 7:
                tid     = _nid(tanq_buffer[0])
                produto = tanq_buffer[1]
                ini     = _fl(tanq_buffer[2])
                fin_raw = _fl(tanq_buffer[6])  # índice 6 = Final (antes de Medição)
                if ini is not None and fin_raw is not None:
                    # Verificar se já existe esse tanque (evitar duplicatas)
                    ids_existentes = [t['id'] for t in tanques]
                    if tid not in ids_existentes:
                        tanques.append({
                            "id":      tid,
                            "produto": produto,
                            "estoque_inicial": ini,
                            "estoque_final":   fin_raw,
                        })
                tanq_buffer = []

    # Fallback: se POSIÇÃO DOS TANQUES não funcionou, tentar POSIÇÃO DOS COMBUSTÍVEIS
    if not tanques:
        em_comb = False
        for l in linhas:
            l = l.strip()
            if re.search(r'POSI[ÇC][ÃA]O\s*DOS\s*COMBUST', l, re.I):
                em_comb = True; continue
            if re.search(r'POSI[ÇC][ÃA]O\s*DOS\s*TANQUE', l, re.I):
                em_comb = False; continue
            if not em_comb: continue
            if re.search(r'^(Produto|Início|Entrada|Venda|Afer|Final|Total|\*)', l, re.I):
                continue
            # Linha de combustível com nome + valores
            partes = [p.strip() for p in re.split(r'\s{2,}|\t', l) if p.strip()]
            if len(partes) >= 7 and not partes[0].isdigit():
                produto = partes[0]
                ini = _fl(partes[1])
                fin = _fl(partes[5]) if len(partes)>5 else None
                if ini and fin:
                    tanques.append({
                        "id": str(len(tanques)+1),
                        "produto": produto,
                        "estoque_inicial": ini,
                        "estoque_final": fin,
                    })

    return {
        "competencia": competencia,
        "tanques": tanques,
        "bicos": bicos,
    }


# ── EXCEL ESTRUTURADO (formato BP Combustíveis e similares) ──────────────────
def _parse_excel_estruturado(arquivo_bytes):
    """Parser para Excel com seções ESTOQUE e MOVIMENTAÇÃO POR BICO."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(arquivo_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    tanques=[]; bicos=[]; competencia=""

    # Detectar período
    for row in rows:
        for cell in row:
            if cell and re.search(r'Periodo|Período', str(cell), re.I):
                # Tentar encontrar datas na mesma linha ou próximas
                for c in row:
                    m = re.search(r'(\d{2}/\d{2}/\d{4})', str(c) if c else "")
                    if m:
                        dt = m.group(1)
                        meses=['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez']
                        try: competencia=f"{meses[int(dt[3:5])-1]}/{dt[6:]}"
                        except: pass

    em_bicos=False; em_tanques=False
    bicos_vistos=set()

    for row in rows:
        c0 = str(row[0]).strip() if row[0] is not None else ""
        c1 = str(row[1]).strip() if len(row)>1 and row[1] is not None else ""

        # Detectar seções
        if re.search(r'MOVIMENTA[ÇC][ÃA]O', c0, re.I):
            em_bicos=True; em_tanques=False; continue
        if re.search(r'^ESTOQUE$', c0, re.I):
            em_tanques=True; em_bicos=False; continue
        if re.search(r'^(Serie|Série|Tanque|Bico|Produto)', c0, re.I):
            continue

        if em_bicos and len(row)>=6:
            # [Serie, Bico, Produto, '', Abertura, Fechamento]
            bid_raw = str(row[1]).strip() if row[1] is not None else ""
            if bid_raw.isdigit() and bid_raw not in bicos_vistos:
                ini = _fl(row[4]) if len(row)>4 and row[4] is not None else None
                fin = _fl(row[5]) if len(row)>5 and row[5] is not None else None
                if ini is not None and fin is not None:
                    bicos.append({"id":_nid(bid_raw),"encerrante_inicial":ini,"encerrante_final":fin})
                    bicos_vistos.add(bid_raw)

        if em_tanques and len(row)>=8:
            # [Tanque, Combustivel, '', '', Est.Abert, '', '', Est.Fech]
            tid_raw = c0
            if tid_raw.isdigit():
                produto = c1
                ini = _fl(row[4]) if row[4] is not None else None
                fin = _fl(row[7]) if len(row)>7 and row[7] is not None else None
                if ini is not None and fin is not None:
                    tanques.append({"id":_nid(tid_raw),"produto":produto,
                                    "estoque_inicial":ini,"estoque_final":fin})

    return {"competencia":competencia,"tanques":tanques,"bicos":bicos}


# ── FUNÇÃO PRINCIPAL ──────────────────────────────────────────────────────────
def ler_dac(arquivo_bytes, filename):
    """
    Lê o DAC em qualquer formato suportado.
    Retorna dict {"competencia", "tanques", "bicos"} ou None se não suportado.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    try:
        if ext in ("xlsx", "xlsm"):
            # Tentar parser estruturado primeiro, depois texto
            resultado = _parse_excel_estruturado(arquivo_bytes)
            if not resultado["tanques"] and not resultado["bicos"]:
                resultado = _ler_excel(arquivo_bytes, filename)
            return resultado if (resultado["tanques"] or resultado["bicos"]) else None

        elif ext == "xls":
            return _ler_excel(arquivo_bytes, filename)

        elif ext == "pdf":
            return _ler_pdf(arquivo_bytes)  # retorna None se escaneado

        else:
            return None  # formato não suportado

    except Exception as e:
        print(f"Erro ao ler DAC ({filename}): {e}")
        return None


# ── CONFRONTO DAC × SPED ──────────────────────────────────────────────────────
def confrontar_dac_sped(dac, d_atu):
    from app import _sk
    resultado = {"tanques": [], "bicos": [], "competencia": dac.get("competencia", "")}

    tanq_pri={}; tanq_ult={}
    for (t,d) in d_atu["tanques"]:
        if t not in tanq_pri or d<tanq_pri[t]: tanq_pri[t]=d
        if t not in tanq_ult or d>tanq_ult[t]: tanq_ult[t]=d

    bico_pri={}; bico_ult={}
    for (b,d) in d_atu["bicos"]:
        if b not in bico_pri or d<bico_pri[b]: bico_pri[b]=d
        if b not in bico_ult or d>bico_ult[b]: bico_ult[b]=d

    def _st(dif):
        if dif is None: return "⚠️ AUSENTE"
        return "✅ OK" if abs(dif) < 0.01 else "❌ DIVERGÊNCIA"

    # Tanques
    for item in sorted(dac.get("tanques",[]), key=lambda x: _sk(str(x.get("id","")))):
        tid     = str(item.get("id","")).strip()
        produto = item.get("produto","")
        ei_dac  = item.get("estoque_inicial")
        ef_dac  = item.get("estoque_final")
        pri_d   = tanq_pri.get(tid); ult_d = tanq_ult.get(tid)
        r_pri   = d_atu["tanques"].get((tid,pri_d),{}) if pri_d else {}
        r_ult   = d_atu["tanques"].get((tid,ult_d),{}) if ult_d else {}
        ei_sped = r_pri.get("est_abert"); ef_sped = r_ult.get("est_fech")
        dif_ini = round(ei_dac-ei_sped,3) if ei_dac is not None and ei_sped is not None else None
        dif_fin = round(ef_dac-ef_sped,3) if ef_dac is not None and ef_sped is not None else None

        # Diagnóstico
        dias_t = sorted(d for (tt,d) in d_atu["tanques"] if tt==tid)
        total_rec  = round(sum(d_atu["tanques"][(tid,d)].get("entrada",0) for d in dias_t),3)
        total_sai  = round(sum(d_atu["tanques"][(tid,d)].get("saida",0)   for d in dias_t),3)
        total_evap = round(sum(d_atu["tanques"][(tid,d)].get("evap",0)    for d in dias_t),3)
        total_aj   = round(sum(d_atu["tanques"][(tid,d)].get("ajuste",0)  for d in dias_t),3)

        diagnostico = ""
        if dif_fin is not None and abs(dif_fin) >= 0.01:
            pct_est = abs(dif_fin)/ef_sped*100 if ef_sped else 0
            pct_ven = abs(dif_fin)/total_sai*100 if total_sai else 0
            if dif_fin > 0:
                causas = []
                if total_aj>0: causas.append(f"ajuste positivo já lançado ({total_aj:,.3f} L)")
                causas += ["entrada não lançada no sistema","venda registrada a maior que a real","ganho por temperatura/dilatação"]
                diagnostico=(f"SOBRA FÍSICA de {dif_fin:,.3f} L ({pct_est:.1f}% do est. final). "
                             f"Possíveis causas: {'; '.join(causas[:3])}.")
            else:
                causas = []
                if total_aj<0: causas.append(f"ajuste negativo já lançado ({total_aj:,.3f} L)")
                causas += [f"venda maior que vol. escritural ({pct_ven:.1f}% das saídas)",
                           f"evaporação não contabilizada (lançado: {total_evap:,.3f} L)",
                           "vazamento ou perda não registrada"]
                diagnostico=(f"FALTA FÍSICA de {abs(dif_fin):,.3f} L ({pct_est:.1f}% do est. final). "
                             f"Possíveis causas: {'; '.join(causas[:3])}.")

        resultado["tanques"].append({
            "id":tid,"produto":produto,
            "ei_dac":ei_dac,"ei_sped":ei_sped,"dif_ini":dif_ini,"status_ini":_st(dif_ini),
            "ef_dac":ef_dac,"ef_sped":ef_sped,"dif_fin":dif_fin,"status_fin":_st(dif_fin),
            "diagnostico":diagnostico,
            "total_entrada":total_rec,"total_saida":total_sai,
            "total_evap":total_evap,"total_ajuste":total_aj,
        })

    # Bicos
    for item in sorted(dac.get("bicos",[]), key=lambda x: _sk(str(x.get("id","")))):
        bid    = str(item.get("id","")).strip()
        ei_dac = item.get("encerrante_inicial"); ef_dac = item.get("encerrante_final")
        pri_d  = bico_pri.get(bid); ult_d = bico_ult.get(bid)
        r_pri  = d_atu["bicos"].get((bid,pri_d),{}) if pri_d else {}
        r_ult  = d_atu["bicos"].get((bid,ult_d),{}) if ult_d else {}
        ei_sped= r_pri.get("enc_abert"); ef_sped = r_ult.get("enc_fech")
        dif_ini= round(ei_dac-ei_sped,3) if ei_dac is not None and ei_sped is not None else None
        dif_fin= round(ef_dac-ef_sped,3) if ef_dac is not None and ef_sped is not None else None
        resultado["bicos"].append({
            "id":bid,
            "ei_dac":ei_dac,"ei_sped":ei_sped,"dif_ini":dif_ini,"status_ini":_st(dif_ini),
            "ef_dac":ef_dac,"ef_sped":ef_sped,"dif_fin":dif_fin,"status_fin":_st(dif_fin),
        })

    return resultado
